#!/usr/bin/env python3
"""
Create MEMA Financial Services Cyber Essentials Question Set Excel
"""

import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "CE Answers"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
section_font = Font(bold=True, size=11)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
ready_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
action_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

headers = ["Section", "Q No.", "Question", "Your Answer", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_alignment
    cell.border = thin_border

# MEMA Financial Services Answers
data = [
    # A1 - Organisation
    ("A1 Organisation", "A1.1", "Organisation name", "MEMA Financial Services Ltd", "Ready"),
    ("A1 Organisation", "A1.2", "Organisation type", "LTD - Limited Company", "Ready"),
    ("A1 Organisation", "A1.3", "Registration number", "15382445", "Ready"),
    ("A1 Organisation", "A1.4", "Registered address", "34-35 Hatton Garden, London, England, EC1N 8DX", "Ready"),
    ("A1 Organisation", "A1.5", "Main business", "IT", "Ready"),
    ("A1 Organisation", "A1.6", "Website address", "https://plannetic.com", "Ready"),
    ("A1 Organisation", "A1.7", "Renewal or first time", "First Time Application", "Ready"),
    ("A1 Organisation", "A1.8", "Two main reasons for applying", "Primary: To give confidence to our customers\nSecondary: To generally improve our security", "Ready"),
    ("A1 Organisation", "A1.9", "Read CE Requirements document?", "Yes", "Ready"),
    ("A1 Organisation", "A1.10", "IASME breach contact permission", "No", "Ready"),
    ("A1 Organisation", "A1.11", "IASME research contact permission", "No", "Ready"),

    # A2 - Scope
    ("A2 Scope", "A2.1", "Whole organisation in scope?", "Yes", "Ready"),
    ("A2 Scope", "A2.2", "Scope description (if not whole org)", "N/A - whole organisation in scope", "N/A"),
    ("A2 Scope", "A2.3", "Geographical locations in scope", "All UK locations. MEMA Financial Services operates fully remotely with 2 home-based employees.", "Ready"),
    ("A2 Scope", "A2.4", "Laptops/desktops (qty + make + OS)", "2 laptops:\n- 1 x Apple MacBook Pro running macOS 26.1\n- 1 x Dell laptop running Windows 11", "Ready"),
    ("A2 Scope", "A2.4.1", "Thin clients", "None", "Ready"),
    ("A2 Scope", "A2.5", "Servers/virtual servers/hypervisors/VDI", "None - all infrastructure is cloud-hosted via AWS (Vercel) and Supabase (PaaS).", "Ready"),
    ("A2 Scope", "A2.6", "Tablets and mobile devices", "1 x Samsung mobile phone running Android 15", "Ready"),
    ("A2 Scope", "A2.7", "Networks in scope", "Home/Remote Worker Networks - both employees work from home using domestic broadband connections. All devices are secured by software firewalls (macOS Application Firewall and Windows Defender Firewall).", "Ready"),
    ("A2 Scope", "A2.7.1", "Remote workers count", "2 - all employees are home/remote workers", "Ready"),
    ("A2 Scope", "A2.8", "Network equipment (firewalls/routers)", "All staff are remote workers using home broadband routers (various makes/models managed by ISPs). Software firewalls are enabled on all endpoint devices. Staff do not have administrative access to home routers.", "Ready"),
    ("A2 Scope", "A2.9", "Cloud services (third party)", "- Supabase (PaaS) - Database and authentication\n- Vercel/AWS (PaaS) - Application hosting\n- GitHub (SaaS) - Source code repository\n- Google Workspace (SaaS) - Email and collaboration\n- 1Password (SaaS) - Password management\n- Resend (SaaS) - Email delivery\n- OpenSign (SaaS) - Document signing\n- Ordnance Survey (SaaS) - UK address lookup API\n- DeepSeek (SaaS) - AI/LLM API\n- Alpha Vantage (SaaS) - Financial data API", "Ready"),
    ("A2 Scope", "A2.10", "Person responsible for IT", "Ade Omosanya, Director", "Ready"),

    # A3 - Insurance
    ("A3 Insurance", "A3.1", "UK HQ and turnover < £20m?", "Yes", "Ready"),
    ("A3 Insurance", "A3.2", "Opt out of insurance?", "No - we wish to receive the included cyber insurance", "Ready"),
    ("A3 Insurance", "A3.3", "Insurance contact email", "info@memaconsultants.com", "Confirm email"),

    # A4 - Firewalls
    ("A4 Firewalls", "A4.1", "Firewalls at internet boundaries?", "Yes", "Ready"),
    ("A4 Firewalls", "A4.1.1", "Software firewalls enabled on all devices?", "Yes - we use the built-in macOS Application Firewall and Windows Defender Firewall; both are enabled and kept enabled at all times, including on untrusted networks (home/public Wi-Fi).", "ACTION: Verify both firewalls are ON"),
    ("A4 Firewalls", "A4.1.2", "If no software firewall: list OS", "N/A", "N/A"),
    ("A4 Firewalls", "A4.2", "Default passwords changed on firewalls?", "Yes - for home workers, domestic routers are managed by ISPs with unique per-device passwords. Software firewalls are protected by device login credentials following our password policy.", "Ready"),
    ("A4 Firewalls", "A4.2.1", "Process for changing firewall password", "Software firewalls are protected by device login credentials. Password changes are performed via System Settings (macOS) or Windows Settings when required. Device passwords follow our policy of minimum 12 characters.", "Ready"),
    ("A4 Firewalls", "A4.3", "Firewall password configuration (A-E)", "Option C - Minimum 12 characters, no maximum length", "Ready"),
    ("A4 Firewalls", "A4.4", "Change password when compromised?", "Yes - if we know or suspect a device password has been compromised, we immediately change it and review for unauthorised access.", "Ready"),
    ("A4 Firewalls", "A4.5", "Process to manage firewall?", "Yes", "Ready"),
    ("A4 Firewalls", "A4.6", "Firewall rules reviewed in last 12 months?", "Software firewalls use default secure settings blocking all incoming connections. No custom inbound rules configured. Configuration verified during device setup and periodically checked. Cloud services (Vercel, Supabase) manage their own infrastructure firewalls.", "Ready"),
    ("A4 Firewalls", "A4.7", "Allows unauthenticated inbound?", "No - all inbound connections blocked by default on endpoint devices.", "Ready"),
    ("A4 Firewalls", "A4.8", "How inbound connections approved?", "No inbound connections permitted to endpoint devices. All services accessed via outbound connections to cloud platforms. Any exception would require documented written approval from the Director with clear business justification.", "Ready"),
    ("A4 Firewalls", "A4.9", "Admin interface accessible from internet?", "No - firewall settings only accessible locally via device System Settings.", "Ready"),
    ("A4 Firewalls", "A4.10", "If yes: documented business need?", "N/A", "N/A"),
    ("A4 Firewalls", "A4.11", "If yes: MFA or IP allow list?", "N/A", "N/A"),

    # A5 - Secure Configuration
    ("A5 Secure Config", "A5.1", "Removed unnecessary software?", "All devices configured with only business-required software. Staff instructed not to install unnecessary applications. macOS Gatekeeper and Windows SmartScreen prevent unsigned/untrusted applications. Regular reviews conducted to remove unused software.", "Ready"),
    ("A5 Secure Config", "A5.2", "Only necessary user accounts?", "Yes - only required user accounts exist on devices and cloud services. Guest accounts disabled. Unused accounts removed promptly.", "Ready"),
    ("A5 Secure Config", "A5.3", "Default passwords changed?", "Yes - all default passwords changed to strong, unique passwords managed via 1Password.", "Ready"),
    ("A5 Secure Config", "A5.4", "Run external services with non-public data?", "Yes - Plannetic is a web application accessed by authenticated IFA clients over the internet. It handles client financial and compliance data.", "Ready"),
    ("A5 Secure Config", "A5.5", "Authentication option (A-E)", "Option A - Multi-factor authentication with minimum 8 character password. Plannetic uses Supabase Auth with MFA (TOTP) for all users.", "Ready"),
    ("A5 Secure Config", "A5.6", "Process to change password on compromise", "1) User notified immediately and required to change password\n2) All sessions terminated via Supabase admin dashboard\n3) MFA verified/re-enrolled\n4) Incident logged and reviewed", "Ready"),
    ("A5 Secure Config", "A5.7", "Brute force protection (A-C)", "Option A - Throttling. Supabase Auth provides rate limiting and account lockout after failed attempts.", "Ready"),
    ("A5 Secure Config", "A5.8", "Auto-run disabled?", "Yes - macOS doesn't auto-run from external media. Windows SmartScreen blocks untrusted apps. Users prompted before opening downloads.", "Ready"),
    ("A5 Secure Config", "A5.9", "Device locking enabled?", "Yes - all devices require authentication to access.", "Ready"),
    ("A5 Secure Config", "A5.10", "Device unlock method", "macOS: Touch ID biometric with password fallback (min 8 chars)\nWindows: Windows Hello or password (min 12 chars)\nAndroid: Fingerprint/PIN (min 6 digits)", "Ready"),

    # A6 - Security Updates
    ("A6 Updates", "A6.1", "All OS supported?", "Yes - all devices run supported operating systems:\n- macOS 26.1 (Tahoe) - supported\n- Windows 11 - supported\n- Android 15 - supported", "Ready"),
    ("A6 Updates", "A6.2", "All software supported?", "Yes - all software licensed and receiving security updates.", "Ready"),
    ("A6 Updates", "A6.2.1", "Browsers (name + version)", "Google Chrome 143.0.7499.110\nMozilla Firefox 146.0\nMicrosoft Edge 143.0.3650.80\nSafari 26.1", "Ready"),
    ("A6 Updates", "A6.2.2", "Malware protection", "macOS: XProtect (built-in, auto-updating) + Gatekeeper + MRT\nWindows: Microsoft Defender Antivirus (built-in, auto-updating)", "Ready"),
    ("A6 Updates", "A6.2.3", "Email applications", "Google Workspace Gmail (web-based via browser)\nApple Mail 16.0 (macOS)", "Ready"),
    ("A6 Updates", "A6.2.4", "Office applications", "Google Workspace (Docs, Sheets, Slides - web-based)\nApple Pages 14.4, Numbers 14.4, Keynote 14.4\nMicrosoft OneNote 16.104", "Ready"),
    ("A6 Updates", "A6.3", "Any unlicensed/unsupported software?", "No", "Ready"),
    ("A6 Updates", "A6.3.1", "List unsupported/unlicensed", "N/A", "N/A"),
    ("A6 Updates", "A6.4", "Critical updates within 14 days?", "Yes - all high/critical security updates applied within 14 days. Automatic updates enabled on all devices.", "Ready"),
    ("A6 Updates", "A6.4.1", "OS auto-updates enabled?", "Yes - automatic updates enabled on both macOS and Windows devices.", "Ready"),
    ("A6 Updates", "A6.4.2", "If not auto: how ensure 14 days?", "Automatic updates are primary mechanism. Director monitors vendor release notes and manually triggers if not auto-applied within 7 days.", "Ready"),
    ("A6 Updates", "A6.5", "App updates within 14 days?", "Yes", "Ready"),
    ("A6 Updates", "A6.5.1", "App auto-updates enabled?", "Yes - App Store (macOS), Microsoft Store (Windows), Chrome/Firefox/Edge all have auto-updates enabled.", "Ready"),
    ("A6 Updates", "A6.5.2", "If not auto: how ensure 14 days?", "Auto-updates primary. Staff accept update prompts promptly. Periodic compliance checks by Director.", "Ready"),
    ("A6 Updates", "A6.6", "Remove unsupported when EOL?", "Yes - unsupported software removed when it reaches end-of-life.", "Ready"),
    ("A6 Updates", "A6.7", "Unsupported moved out of scope?", "Not applicable - no unsupported software used.", "Ready"),

    # A7 - User Access Control
    ("A7 Access Control", "A7.1", "Account creation approval process", "As a 2-person company, the Director (Ade Omosanya) approves all account creation. New accounts created with minimum necessary permissions. MFA enrollment mandatory before first access.", "Ready"),
    ("A7 Access Control", "A7.2", "Unique credentials?", "Yes - all users have unique credentials. Account sharing prohibited.", "Ready"),
    ("A7 Access Control", "A7.3", "Leaver process", "When an employee leaves: 1) All cloud service accounts immediately disabled/deleted 2) Device collected and wiped 3) Passwords to shared systems changed 4) Confirmed within 24 hours", "Ready"),
    ("A7 Access Control", "A7.4", "Least privilege", "Role-based access control (RBAC) implemented in Plannetic. Users granted minimum permissions for their role. Cloud service access limited to job requirements.", "Ready"),
    ("A7 Access Control", "A7.5", "Admin access process", "Director approval required for admin access. Separate admin accounts used. Access documented and reviewed quarterly.", "Ready"),
    ("A7 Access Control", "A7.6", "Separate admin accounts", "Admin tasks performed using dedicated accounts. Cloud platforms: separate admin accounts with MFA. Day-to-day work uses standard accounts.", "Ready"),
    ("A7 Access Control", "A7.7", "Prevent admin for email/browsing", "Admin accounts not used for email or browsing. Staff trained on this policy. Cloud admin sessions logged out after use.", "Ready"),
    ("A7 Access Control", "A7.8", "Track admin accounts?", "Yes - Director maintains record of all admin access.", "Ready"),
    ("A7 Access Control", "A7.9", "Review admin access regularly?", "Yes - reviewed quarterly. Unused privileges removed.", "Ready"),
    ("A7 Access Control", "A7.10", "Brute-force protection", "All systems protected:\n- 1Password: Account lockout\n- Google Workspace: Rate limiting + MFA\n- Supabase: Rate limiting + lockout\n- Device logins: Throttling after failed attempts", "Ready"),
    ("A7 Access Control", "A7.11", "Password quality controls", "Technical controls via 1Password and cloud providers:\n- Minimum 8 characters with MFA enabled\n- Minimum 12 characters where MFA not available\n- No maximum length\n- Common password blocking", "Ready"),
    ("A7 Access Control", "A7.12", "Support strong passwords", "1Password used for generating and storing unique passwords. Staff trained on three random words method. No regular password expiry. No complexity requirements.", "Ready"),
    ("A7 Access Control", "A7.13", "Compromised password process?", "Yes - report immediately, force reset, terminate sessions, re-verify MFA, review incident.", "Ready"),
    ("A7 Access Control", "A7.14", "All cloud services have MFA?", "Yes - all cloud services provide MFA:\n- Google Workspace ✓\n- Supabase ✓\n- Vercel ✓\n- GitHub ✓\n- 1Password ✓", "Ready"),
    ("A7 Access Control", "A7.15", "Services without MFA option", "N/A - all services have MFA", "N/A"),
    ("A7 Access Control", "A7.16", "MFA on all cloud admins?", "Yes - MFA mandatory for all admin accounts.", "Ready"),
    ("A7 Access Control", "A7.17", "MFA on all cloud users?", "Yes - MFA mandatory for all users.", "Ready"),

    # A8 - Malware Protection
    ("A8 Malware", "A8.1", "Malware protection method (A/B/C)", "Both A and B:\nA - Anti-malware: Windows Defender, macOS XProtect\nB - Allow-listing: Gatekeeper (macOS), SmartScreen (Windows), Google Play Protect (Android)", "Ready"),
    ("A8 Malware", "A8.2", "Anti-malware updates + blocks?", "Yes - Windows Defender and XProtect auto-update and block/quarantine detected malware.", "Ready"),
    ("A8 Malware", "A8.3", "Warns on malicious sites?", "Yes:\n- Chrome: Safe Browsing enabled\n- Firefox: Phishing protection enabled\n- Edge: SmartScreen enabled\n- Safari: Fraudulent Website Warning enabled", "Ready"),
    ("A8 Malware", "A8.4", "Users restricted from unsigned apps?", "Yes - Gatekeeper (macOS) and SmartScreen (Windows) block unsigned/untrusted apps. Users cannot bypass without admin intervention.", "Ready"),
    ("A8 Malware", "A8.5", "Only approved apps + list maintained?", "Yes - staff instructed to only install approved business applications. Director approves new app requests. Mobile: only official app stores used.", "Ready"),
]

# Write data
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        if col_num == 1:
            cell.font = section_font
        if col_num == 5:
            if value == "Ready":
                cell.fill = ready_fill
            elif "ACTION" in value or "Confirm" in value:
                cell.fill = action_fill

# Column widths
widths = [18, 8, 45, 70, 25]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = "A2"

# Summary sheet
ws2 = wb.create_sheet("Actions Required")
ws2["A1"] = "MEMA Financial Services - Cyber Essentials Actions"
ws2["A1"].font = Font(bold=True, size=14)

ws2["A3"] = "Before Submission - Verify:"
ws2["A3"].font = Font(bold=True)

actions = [
    "1. macOS Firewall is ENABLED (System Settings > Network > Firewall)",
    "2. Windows Defender Firewall is ENABLED (Settings > Privacy & security > Windows Security > Firewall)",
    "3. Confirm insurance email address is correct",
    "4. All devices have auto-updates enabled",
    "5. MFA is enforced on all cloud services",
]

for i, action in enumerate(actions, 4):
    ws2[f"A{i}"] = action

ws2["A10"] = "Company Details Confirmed:"
ws2["A10"].font = Font(bold=True)
ws2["A11"] = "MEMA Financial Services Ltd"
ws2["A12"] = "Company Number: 15382445"
ws2["A13"] = "34-35 Hatton Garden, London, EC1N 8DX"
ws2["A14"] = "IT Contact: Ade Omosanya, Director"

ws2.column_dimensions["A"].width = 80

# Save
output = "/Users/adeomosanya/Documents/ifa-professional-portal/cyber-essentials/MEMA-Cyber-Essentials-Answers.xlsx"
wb.save(output)
print(f"Created: {output}")
