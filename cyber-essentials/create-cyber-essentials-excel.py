#!/usr/bin/env python3
"""
Create Cyber Essentials Question Set Excel file with draft answers
"""

import subprocess
import sys

# Install openpyxl if not present
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "CE Question Set"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
section_font = Font(bold=True, size=11)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Section", "Q No.", "Question", "Guidance", "Answer Type", "Draft Answer", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_alignment
    cell.border = thin_border

# Data - Questions and Answers
data = [
    # A1 - Organisation
    ("A1 - Organisation", "A1.1", "What is your organisation's name?", "Name displayed on certificate (max 150 chars)", "Notes", "TBC", "Pending"),
    ("A1 - Organisation", "A1.2", "What type of organisation are you?", "LTD/LLP/CIC/COP/MTL/CHA/GOV/SOL/PRT/SOC/OTH", "Multiple choice", "TBC", "Pending"),
    ("A1 - Organisation", "A1.3", "What is your organisation's registration number?", "Companies House number or 'none'", "Notes", "TBC", "Pending"),
    ("A1 - Organisation", "A1.4", "What is your organisation's address?", "Legal registered address", "Notes", "TBC", "Pending"),
    ("A1 - Organisation", "A1.5", "What is your main business?", "Select from list - likely Finance or IT", "Multiple choice", "TBC", "Pending"),
    ("A1 - Organisation", "A1.6", "What is your website address?", "Website or LinkedIn/Facebook page", "Notes", "https://plannetic.com", "Confirm"),
    ("A1 - Organisation", "A1.7", "Is this application a renewal or first time?", "Renewal or First Time Application", "Select", "First Time Application", "Confirm"),
    ("A1 - Organisation", "A1.8", "What are the two main reasons for applying?", "Select two reasons", "Multiple choice", "TBC - likely 'to give confidence to our customers' + 'to generally improve our security'", "Pending"),
    ("A1 - Organisation", "A1.9", "Have you read the CE Requirements document?", "", "Yes/No", "Yes", "Ready"),
    ("A1 - Organisation", "A1.10", "Can IASME contact you if you experience a breach?", "", "Yes/No", "TBC", "Pending"),
    ("A1 - Organisation", "A1.11", "Can IASME contact you for research purposes?", "", "Yes/No", "TBC", "Pending"),

    # A2 - Scope
    ("A2 - Scope", "A2.1", "Does the scope cover your whole organisation?", "Whole org = eligible for free insurance", "Yes/No", "Yes", "Confirm"),
    ("A2 - Scope", "A2.2", "Scope description (if not whole org)", "Only if A2.1 = No", "Notes", "N/A", "N/A"),
    ("A2 - Scope", "A2.3", "Geographical locations in scope", "List locations or 'All UK offices'", "Notes", "All UK locations including home/remote working locations for all staff.", "Ready"),
    ("A2 - Scope", "A2.4", "Laptops/desktops/virtual desktops (qty + make + OS)", "Must include make and OS version", "Notes", "1 x Apple Mac laptop running macOS 26.1 (Build 25B78)", "Ready - add more if needed"),
    ("A2 - Scope", "A2.4.1", "Thin clients (qty + make + OS)", "", "Notes", "None", "Ready"),
    ("A2 - Scope", "A2.5", "Servers/virtual servers/hypervisors/VDI", "", "Notes", "None - all infrastructure is cloud-hosted via Vercel (PaaS) and Supabase (PaaS).", "Ready"),
    ("A2 - Scope", "A2.6", "Tablets and mobile devices (qty + make + OS)", "", "Notes", "TBC", "Pending"),
    ("A2 - Scope", "A2.7", "Networks in scope (name + location + purpose)", "", "Notes", "Home/Remote Worker Networks - staff work from home using domestic broadband connections. All devices secured by software firewalls.", "Ready"),
    ("A2 - Scope", "A2.7.1", "How many staff are home or remote workers?", "", "Notes", "TBC", "Pending"),
    ("A2 - Scope", "A2.8", "Network equipment in scope (firewalls/routers - make + model)", "Must include make and model", "Notes", "All staff are remote workers using home broadband routers (various makes/models managed by ISPs). Software firewalls enabled on all endpoint devices. Staff do not have admin access to home routers.", "Ready"),
    ("A2 - Scope", "A2.9", "Cloud services (third party - SaaS/PaaS/IaaS)", "", "Notes", "Supabase (PaaS), Vercel (PaaS), GitHub (SaaS), Resend (SaaS), OpenSign (SaaS), Ordnance Survey (SaaS), DeepSeek (SaaS), Alpha Vantage (SaaS). [TBC: Add email/collab suite, password manager]", "Partial"),
    ("A2 - Scope", "A2.10", "Person responsible for managing IT (name + role)", "Must be org employee, not outsourced IT", "Notes", "TBC", "Pending"),

    # A3 - Insurance
    ("A3 - Insurance", "A3.1", "UK/Crown Dependencies HQ and turnover < £20m?", "Eligibility for free insurance", "Yes/No", "Yes", "Confirm"),
    ("A3 - Insurance", "A3.2", "Opt out of insurance? (if A3.1 = Yes)", "", "Yes/No", "No - we wish to receive the included cyber insurance", "Ready"),
    ("A3 - Insurance", "A3.3", "Insurance contact email (if not opted out)", "", "Notes", "TBC", "Pending"),

    # A4 - Firewalls
    ("A4 - Firewalls", "A4.1", "Do you have firewalls at internet boundaries?", "", "Yes/No", "Yes", "Ready"),
    ("A4 - Firewalls", "A4.1.1", "Software firewalls enabled on all computers/laptops/servers?", "", "Yes/No", "Yes - we use the built-in macOS Application Firewall; it is enabled and kept enabled at all times, including on untrusted networks (home/public Wi-Fi).", "Ready"),
    ("A4 - Firewalls", "A4.1.2", "If A4.1.1 = No: list OS without firewall", "", "Notes", "N/A", "N/A"),
    ("A4 - Firewalls", "A4.2", "Default admin passwords changed on boundary firewalls?", "", "Yes/No", "Yes - for home workers, domestic routers managed by ISPs with unique passwords. Software firewalls protected by device credentials.", "Ready"),
    ("A4 - Firewalls", "A4.2.1", "Process for changing firewall password", "", "Notes", "Software firewalls protected by device login credentials. Password changes via System Settings > Users & Groups. Follows password policy (min 12 chars or MFA + 8 chars).", "Ready"),
    ("A4 - Firewalls", "A4.3", "Firewall password configuration option (A-E)", "A=MFA+8, B=Block common+8, C=12 chars, D=Passwordless, E=Other", "Select", "Option C - Minimum 12 characters, no maximum length", "Ready"),
    ("A4 - Firewalls", "A4.4", "Change firewall password when compromised?", "", "Yes/No", "Yes", "Ready"),
    ("A4 - Firewalls", "A4.5", "Process to manage firewall?", "", "Yes/No", "Yes", "Ready"),
    ("A4 - Firewalls", "A4.6", "Firewall rules reviewed in last 12 months?", "", "Notes", "Software firewalls use default secure settings blocking all incoming connections. No custom inbound rules. Verified during device setup and periodically checked. Cloud services manage own infrastructure firewalls.", "Ready"),
    ("A4 - Firewalls", "A4.7", "Firewall allows unauthenticated inbound connections?", "", "Yes/No", "No - all inbound blocked by default", "Ready"),
    ("A4 - Firewalls", "A4.8", "How inbound connections approved & documented", "", "Notes", "No inbound connections permitted to endpoints. All services via outbound to cloud. Any exception requires documented written approval from IT manager with business justification.", "Ready"),
    ("A4 - Firewalls", "A4.9", "Firewall admin interface accessible from internet?", "", "Yes/No", "No - only accessible locally via System Settings", "Ready"),
    ("A4 - Firewalls", "A4.10", "If A4.9 = Yes: documented business requirement?", "", "Yes/No", "N/A", "N/A"),
    ("A4 - Firewalls", "A4.11", "If A4.9 = Yes: MFA or IP allow list protection?", "", "Notes", "N/A", "N/A"),

    # A5 - Secure Configuration
    ("A5 - Secure Config", "A5.1", "Removed/disabled unnecessary software? (describe)", "", "Notes", "All devices configured with only business-required software. Staff instructed not to install unnecessary apps. macOS Gatekeeper prevents unsigned apps. Regular reviews to remove unused software.", "Ready"),
    ("A5 - Secure Config", "A5.2", "Only necessary user accounts exist?", "", "Yes/No", "Yes - only required accounts. Guest accounts disabled. Unused accounts removed promptly.", "Ready"),
    ("A5 - Secure Config", "A5.3", "Default/guessable passwords changed?", "", "Yes/No", "Yes - all default passwords changed to strong, unique passwords per password policy.", "Ready"),
    ("A5 - Secure Config", "A5.4", "Run/host external services with non-public data?", "", "Yes/No", "Yes - Plannetic platform accessed by authenticated users over internet, handles client financial data.", "Ready"),
    ("A5 - Secure Config", "A5.5", "If A5.4 = Yes: authentication option (A-E)", "", "Select", "Option A - MFA with min 8 char password. Plannetic uses Supabase Auth with MFA (TOTP).", "Ready"),
    ("A5 - Secure Config", "A5.6", "If A5.4 uses passwords: process to change on compromise", "", "Notes", "1) User notified, must change password; 2) Sessions terminated via Supabase admin; 3) MFA verified/re-enrolled; 4) Incident logged and reviewed.", "Ready"),
    ("A5 - Secure Config", "A5.7", "If A5.4 = Yes and not MFA: brute force protection (A-C)", "", "Select", "Option A - Throttling. Supabase Auth rate limiting + account lockout after failed attempts.", "Ready"),
    ("A5 - Secure Config", "A5.8", "Auto-run/execute disabled?", "", "Yes/No", "Yes - macOS doesn't auto-run from external media. Users prompted before opening downloads. Gatekeeper prevents unsigned auto-run.", "Ready"),
    ("A5 - Secure Config", "A5.9", "Device locking enabled?", "", "Yes/No", "Yes - all devices require authentication", "Ready"),
    ("A5 - Secure Config", "A5.10", "If A5.9 = Yes: device unlock method(s)", "", "Notes", "macOS: Touch ID biometric with password fallback (min 8 chars). Auto-lock after inactivity. Password/biometric on wake.", "Ready"),

    # A6 - Security Updates
    ("A6 - Updates", "A6.1", "All OS supported and receiving updates?", "", "Yes/No", "Yes - macOS 26.1 (Tahoe) fully supported by Apple.", "Ready"),
    ("A6 - Updates", "A6.2", "All software supported with vulnerability fixes?", "", "Yes/No", "Yes - all software licensed and receiving updates.", "Ready"),
    ("A6 - Updates", "A6.2.1", "Browsers (name + version)", "", "Notes", "Google Chrome 143.0.7499.110, Mozilla Firefox 146.0, Microsoft Edge 143.0.3650.80, Safari 26.1", "Ready"),
    ("A6 - Updates", "A6.2.2", "Malware protection software (name + version)", "", "Notes", "macOS: XProtect (built-in, auto-updating) + Gatekeeper + MRT", "Ready"),
    ("A6 - Updates", "A6.2.3", "Email applications (name + version)", "", "Notes", "Apple Mail 16.0, web-based email via browser", "Ready"),
    ("A6 - Updates", "A6.2.4", "Office applications (name + version)", "", "Notes", "Apple Pages 14.4, Apple Numbers 14.4, Apple Keynote 14.4, Microsoft OneNote 16.104", "Ready"),
    ("A6 - Updates", "A6.3", "Any unlicensed/unsupported software?", "", "Yes/No", "No", "Ready"),
    ("A6 - Updates", "A6.3.1", "If A6.3 = Yes: list unsupported/unlicensed", "", "Notes", "N/A", "N/A"),
    ("A6 - Updates", "A6.4", "Critical OS/firmware updates within 14 days?", "", "Yes/No", "Yes - all high/critical updates applied within 14 days. Auto-updates enabled.", "Ready"),
    ("A6 - Updates", "A6.4.1", "OS updates via auto-updates?", "", "Yes/No", "Yes - macOS automatic updates enabled", "Ready"),
    ("A6 - Updates", "A6.4.2", "If not auto: how ensure within 14 days", "", "Notes", "Auto-updates primary mechanism. Monitor vendor release notes, manually trigger if not auto-applied within 7 days.", "Ready"),
    ("A6 - Updates", "A6.5", "Critical application updates within 14 days?", "", "Yes/No", "Yes", "Ready"),
    ("A6 - Updates", "A6.5.1", "App updates via auto-updates?", "", "Yes/No", "Yes - App Store auto-updates, browser auto-updates, Microsoft AutoUpdate", "Ready"),
    ("A6 - Updates", "A6.5.2", "If not auto: how ensure within 14 days", "", "Notes", "Auto-updates primary. Staff trained to accept prompts. Periodic compliance checks.", "Ready"),
    ("A6 - Updates", "A6.6", "Remove unsupported software when EOL?", "", "Yes/No", "Yes", "Ready"),
    ("A6 - Updates", "A6.7", "If unsupported needed: moved out of scope?", "", "Notes", "Not applicable - no unsupported software used.", "Ready"),

    # A7 - User Access Control
    ("A7 - Access Control", "A7.1", "User account creation approval process", "", "Notes", "1) Manager requests; 2) IT verifies & determines role; 3) Account created with min permissions; 4) User receives password setup link; 5) MFA mandatory before first access.", "Ready"),
    ("A7 - Access Control", "A7.2", "Unique credentials for all accounts?", "", "Yes/No", "Yes - all users have unique credentials. Sharing prohibited.", "Ready"),
    ("A7 - Access Control", "A7.3", "Leaver process (disable/delete accounts)", "", "Notes", "1) HR notifies IT of leaving date; 2) On leaving: accounts disabled, access revoked, device wiped; 3) Confirmed within 24 hours.", "Ready"),
    ("A7 - Access Control", "A7.4", "Least privilege process", "", "Notes", "RBAC in Plannetic: Admin, Advisor, Compliance, Support, Client roles. Access reviewed on role change. Periodic access reviews.", "Ready"),
    ("A7 - Access Control", "A7.5", "Admin access granting process", "", "Notes", "1) Written request with justification; 2) Director approval; 3) Separate admin account; 4) Documented in register; 5) Quarterly review.", "Ready"),
    ("A7 - Access Control", "A7.6", "Separate admin accounts for admin tasks", "", "Notes", "Admin tasks use dedicated accounts. Cloud: separate admin with MFA. Devices: standard user for daily work, separate admin for config.", "Ready"),
    ("A7 - Access Control", "A7.7", "Prevent admin accounts for web/email", "", "Notes", "Admin accounts not configured with email. Training that admin not for browsing. Cloud admin logged out after use.", "Ready"),
    ("A7 - Access Control", "A7.8", "Track who has admin accounts?", "", "Yes/No", "Yes - admin access register maintained", "Ready"),
    ("A7 - Access Control", "A7.9", "Review admin access regularly?", "", "Yes/No", "Yes - quarterly review. Unused privileges removed.", "Ready"),
    ("A7 - Access Control", "A7.10", "Brute-force protection for passwords", "", "Notes", "Supabase: rate limiting + lockout. macOS: throttling. Cloud services: vendor protection + MFA.", "Ready"),
    ("A7 - Access Control", "A7.11", "Technical controls for password quality", "", "Notes", "Min 8 chars with MFA (Supabase). Min 12 chars without MFA. No max length. Common password blocking where supported.", "Ready"),
    ("A7 - Access Control", "A7.12", "How support unique/strong passwords", "", "Notes", "Password manager recommended. Three random words training. No regular expiry. No complexity requirements. No reuse training.", "Ready"),
    ("A7 - Access Control", "A7.13", "Compromised password/account process?", "", "Yes/No", "Yes - report, force reset, terminate sessions, re-verify MFA, review incident.", "Ready"),
    ("A7 - Access Control", "A7.14", "All cloud services have MFA available?", "", "Yes/No", "Yes - Supabase, Vercel, GitHub, Resend all have MFA. [Confirm: OpenSign, Alpha Vantage, Ordnance Survey]", "Partial"),
    ("A7 - Access Control", "A7.15", "If A7.14 = No: list services without MFA", "", "Notes", "N/A", "N/A"),
    ("A7 - Access Control", "A7.16", "MFA applied to all cloud admins?", "", "Yes/No", "Yes - MFA mandatory for all admin accounts", "Ready"),
    ("A7 - Access Control", "A7.17", "MFA applied to all cloud users?", "", "Yes/No", "Yes - MFA mandatory for all users", "Ready"),

    # A8 - Malware Protection
    ("A8 - Malware", "A8.1", "Malware protection method(s) (A/B/C)", "A=Anti-malware, B=Allow-listing, C=Other", "Select", "Both A and B: Anti-malware (XProtect, Gatekeeper) + Allow-listing (Gatekeeper code signing, App Store)", "Ready"),
    ("A8 - Malware", "A8.2", "If Option A: anti-malware updates + blocks malware?", "", "Yes/No", "Yes - XProtect auto-updates. Gatekeeper prevents execution. MRT removes detected malware.", "Ready"),
    ("A8 - Malware", "A8.3", "If Option A: scans/warns on malicious sites?", "", "Yes/No", "Yes - Safari Fraudulent Warning, Chrome Safe Browsing, Firefox protection, Edge SmartScreen", "Ready"),
    ("A8 - Malware", "A8.4", "If Option B: users restricted from unsigned apps?", "", "Yes/No", "Yes - Gatekeeper blocks unsigned apps. Cannot run without admin intervention.", "Ready"),
    ("A8 - Malware", "A8.5", "If Option B: only approved apps + list maintained?", "", "Yes/No", "Yes - staff instructed on approved apps. List maintained. New requests need approval.", "Ready"),
]

# Write data
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = wrap_alignment
        cell.border = thin_border

        # Apply section styling
        if col_num == 1:
            cell.font = section_font

        # Color code status
        if col_num == 7:  # Status column
            if value == "Ready":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif value == "Pending":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif value == "Confirm":
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            elif value == "Partial":
                cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Set column widths
column_widths = [18, 8, 50, 40, 15, 60, 12]
for col_num, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Cyber Essentials Certification - Status Summary"
ws2["A1"].font = Font(bold=True, size=14)

ws2["A3"] = "Status Legend:"
ws2["A4"] = "Ready"
ws2["A4"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ws2["B4"] = "Answer complete and verified"

ws2["A5"] = "Confirm"
ws2["A5"].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
ws2["B5"] = "Needs your confirmation"

ws2["A6"] = "Pending"
ws2["A6"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ws2["B6"] = "Requires your input (TBC)"

ws2["A7"] = "Partial"
ws2["A7"].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ws2["B7"] = "Partially complete, needs review"

ws2["A9"] = "Items Requiring Your Input:"
ws2["A9"].font = Font(bold=True)

pending_items = [
    "A1.1 - Organisation name",
    "A1.2 - Organisation type",
    "A1.3 - Registration number",
    "A1.4 - Registered address",
    "A1.5 - Main business",
    "A1.8 - Reasons for certification",
    "A1.10/A1.11 - Contact permissions",
    "A2.6 - Mobile devices",
    "A2.7.1 - Remote workers count",
    "A2.10 - IT responsible person",
    "A3.3 - Insurance contact email",
]

for i, item in enumerate(pending_items, 10):
    ws2[f"A{i}"] = item

ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 40

# Save
output_path = "/Users/adeomosanya/Documents/ifa-professional-portal/cyber-essentials/Cyber-Essentials-Question-Set-Answers.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
