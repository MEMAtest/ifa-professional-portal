#!/usr/bin/env python3
"""
Plannetic Pricing Analysis v3 - Excel Generator with Charts
Added: Summary chart, Revenue Calculator chart, improved ROI chart
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

# Create workbook
wb = Workbook()

# ============================================
# STYLES
# ============================================
title_font = Font(name='Calibri', size=20, bold=True, color='1E40AF')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
subheader_font = Font(name='Calibri', size=13, bold=True, color='1E40AF')
section_font = Font(name='Calibri', size=11, bold=True, color='374151')
normal_font = Font(name='Calibri', size=10)
small_font = Font(name='Calibri', size=9, color='6B7280')

header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
alt_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
highlight_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
input_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# ============================================
# SHEET 1: EXECUTIVE SUMMARY (with Chart)
# ============================================
ws_summary = wb.active
ws_summary.title = "Summary"

ws_summary['A1'] = "PLANNETIC PRICING ANALYSIS"
ws_summary['A1'].font = title_font
ws_summary.merge_cells('A1:F1')

ws_summary['A2'] = "Verified December 2025"
ws_summary['A2'].font = small_font

ws_summary['A4'] = "What is Plannetic?"
ws_summary['A4'].font = subheader_font

summary_text = [
    "Plannetic is a comprehensive, compliance-focused financial advisory platform",
    "designed specifically for UK-regulated Independent Financial Advisors (IFAs).",
    "",
    "Key Value Proposition:",
    "• Replaces 5-7 separate tools with one integrated platform",
    "• Saves IFAs £280+/month vs competitor tool stack (£530 → £250)",
    "• Saves 10-20 hours per client onboarding",
    "• Built-in FCA compliance and Consumer Duty workflows",
]

for i, text in enumerate(summary_text, start=6):
    ws_summary[f'A{i}'] = text
    ws_summary[f'A{i}'].font = normal_font

# Cost Comparison Data for Chart (visible to user)
ws_summary['A16'] = "Cost Comparison"
ws_summary['A16'].font = subheader_font

cost_headers = ['Category', 'Cost (£/mo)']
for col, header in enumerate(cost_headers, start=1):
    cell = ws_summary.cell(row=18, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

cost_data = [
    ['Competitor Stack', 530],
    ['Plannetic Standard', 250],
    ['Plannetic Professional', 300],
    ['Monthly Savings', 280],
]

for row_idx, (label, value) in enumerate(cost_data, start=19):
    ws_summary.cell(row=row_idx, column=1, value=label).border = thin_border
    cell = ws_summary.cell(row=row_idx, column=2, value=value)
    cell.number_format = '£#,##0'
    cell.border = thin_border
    if row_idx == 22:  # Savings row
        ws_summary.cell(row=row_idx, column=1).fill = highlight_fill
        cell.fill = highlight_fill

# Add bar chart comparing costs
chart_summary = BarChart()
chart_summary.type = "col"
chart_summary.style = 10
chart_summary.title = "Monthly Cost: Competitors vs Plannetic"
chart_summary.y_axis.title = "£ per month"
chart_summary.y_axis.numFmt = '£#,##0'

data_summary = Reference(ws_summary, min_col=2, min_row=18, max_row=21)  # Exclude savings row
cats_summary = Reference(ws_summary, min_col=1, min_row=19, max_row=21)
chart_summary.add_data(data_summary, titles_from_data=True)
chart_summary.set_categories(cats_summary)
chart_summary.width = 12
chart_summary.height = 8
chart_summary.legend = None

ws_summary.add_chart(chart_summary, "D16")

# Pricing Tiers Table
ws_summary['A25'] = "Recommended Pricing Tiers"
ws_summary['A25'].font = subheader_font

pricing_headers = ['Tier', 'Monthly', 'Commitment', '2-Year TCV', '3-Year TCV', 'Best For']
for col, header in enumerate(pricing_headers, start=1):
    cell = ws_summary.cell(row=27, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

pricing_data = [
    ['Monthly', 350, 'Month-to-month', '=B28*24', '=B28*36', 'Trial/uncertain firms'],
    ['Standard', 250, '2-year', '=B29*24', '=B29*36', 'Solo advisors, small firms'],
    ['Professional', 300, '2-year', '=B30*24', '=B30*36', 'Growing firms, AI + support'],
    ['Enterprise', 'Custom', '3-year', 'Custom', 'Custom', '5+ advisors, white-label'],
]

for row_idx, row_data in enumerate(pricing_data, start=28):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if row_idx % 2 == 1:
            cell.fill = alt_fill
        if col_idx == 2 and isinstance(value, int):
            cell.number_format = '£#,##0'
        if col_idx in [4, 5] and isinstance(value, str) and value.startswith('='):
            cell.number_format = '£#,##0'

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 12
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 14
ws_summary.column_dimensions['E'].width = 14
ws_summary.column_dimensions['F'].width = 26

# ============================================
# SHEET 2: COMPETITOR ANALYSIS (with Charts)
# ============================================
ws_comp = wb.create_sheet("Competitor Pricing")

ws_comp['A1'] = "COMPETITOR PRICING ANALYSIS"
ws_comp['A1'].font = title_font
ws_comp.merge_cells('A1:E1')

ws_comp['A2'] = "Verified December 2025 - Sources linked below"
ws_comp['A2'].font = small_font

ws_comp['A4'] = "UK IFA Software Market - Verified Pricing"
ws_comp['A4'].font = subheader_font

comp_headers = ['Software', 'Monthly Price', 'Price Type', 'What They Offer', 'Source']
for col, header in enumerate(comp_headers, start=1):
    cell = ws_comp.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

competitors = [
    ['Intelliflo Office', 132, 'Per user', 'Back office + cashflow', 'TrustRadius'],
    ['Voyant AdviserGo', 175, 'Flat fee', 'Cash flow planning', 'voyant.com'],
    ['Timeline', 162, 'Flat (+VAT)', 'Monte Carlo simulations', 'timeline.co'],
    ['FE CashCalc', 90, 'Per adviser (+VAT)', 'Cash flow modelling', 'advisoryai.com'],
    ['Dynamic Planner', 200, 'Estimated', 'Risk profiling + reports', 'Contact required'],
    ['Plannetic Standard', 250, 'Flat fee', 'ALL-IN-ONE PLATFORM', 'Your price'],
]

for row_idx, row_data in enumerate(competitors, start=7):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx == 2:
            cell.number_format = '£#,##0'
        if row_idx == 12:
            cell.fill = highlight_fill
            cell.font = Font(name='Calibri', size=10, bold=True)
        elif row_idx % 2 == 0:
            cell.fill = alt_fill

# Bar chart for competitor pricing
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Monthly Software Costs Comparison"
chart1.y_axis.title = "£ per month"
chart1.y_axis.numFmt = '£#,##0'

data = Reference(ws_comp, min_col=2, min_row=6, max_row=12)
cats = Reference(ws_comp, min_col=1, min_row=7, max_row=12)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.width = 14
chart1.height = 10
chart1.legend = None

ws_comp.add_chart(chart1, "A15")

# Tool Stack Comparison
ws_comp['A32'] = "Typical IFA Tool Stack vs Plannetic"
ws_comp['A32'].font = subheader_font

stack_headers = ['Tool Category', 'Standalone Cost', 'Plannetic', 'Savings']
for col, header in enumerate(stack_headers, start=1):
    cell = ws_comp.cell(row=34, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

stack_data = [
    ['CRM (generic)', 50, 'Included', '=B35'],
    ['Risk Profiling', 50, 'Included', '=B36'],
    ['Cash Flow (Voyant)', 175, 'Included', '=B37'],
    ['Monte Carlo (Timeline)', 162, 'Included', '=B38'],
    ['Document Generation', 50, 'Included', '=B39'],
    ['E-Signatures', 23, 'Included', '=B40'],
    ['Compliance Tracking', 50, 'Included', '=B41'],
]

for row_idx, row_data in enumerate(stack_data, start=35):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx in [2, 4]:
            cell.number_format = '£#,##0'

# Total row
ws_comp['A42'] = 'TOTAL'
ws_comp['B42'] = '=SUM(B35:B41)'
ws_comp['C42'] = '£250/mo'
ws_comp['D42'] = '=B42-250'

for col in range(1, 5):
    cell = ws_comp.cell(row=42, column=col)
    cell.font = Font(name='Calibri', size=10, bold=True)
    cell.fill = highlight_fill
    cell.border = thin_border
    if col in [2, 4]:
        cell.number_format = '£#,##0'

# Pie chart for tool stack
chart2 = PieChart()
chart2.title = "Tool Stack Cost Breakdown"
data2 = Reference(ws_comp, min_col=2, min_row=35, max_row=41)
labels2 = Reference(ws_comp, min_col=1, min_row=35, max_row=41)
chart2.add_data(data2)
chart2.set_categories(labels2)
chart2.width = 11
chart2.height = 9

chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.dataLabels.showVal = False
chart2.dataLabels.showCatName = False

ws_comp.add_chart(chart2, "F32")

# Sources
ws_comp['A45'] = "Sources:"
ws_comp['A45'].font = section_font
sources = [
    "• Voyant: planwithvoyant.com/uk/pricing",
    "• Timeline: timeline.co (£135+VAT = £162)",
    "• CashCalc: advisoryai.com (£75+VAT = £90)",
    "• Intelliflo: trustradius.com (£130-135/user)",
]
for i, source in enumerate(sources, start=46):
    ws_comp[f'A{i}'] = source
    ws_comp[f'A{i}'].font = small_font

# Column widths
ws_comp.column_dimensions['A'].width = 22
ws_comp.column_dimensions['B'].width = 15
ws_comp.column_dimensions['C'].width = 18
ws_comp.column_dimensions['D'].width = 12
ws_comp.column_dimensions['E'].width = 18

# ============================================
# SHEET 3: REVENUE CALCULATOR (with Chart)
# ============================================
ws_calc = wb.create_sheet("Revenue Calculator")

ws_calc['A1'] = "REVENUE CALCULATOR"
ws_calc['A1'].font = title_font
ws_calc.merge_cells('A1:G1')

# Input Section
ws_calc['A3'] = "INPUT PARAMETERS (Edit yellow cells)"
ws_calc['A3'].font = subheader_font

ws_calc['A5'] = "Standard Monthly Rate (£)"
ws_calc['B5'] = 250
ws_calc['B5'].number_format = '£#,##0'
ws_calc['B5'].fill = input_fill
ws_calc['B5'].border = thin_border

ws_calc['A6'] = "Professional Monthly Rate (£)"
ws_calc['B6'] = 300
ws_calc['B6'].number_format = '£#,##0'
ws_calc['B6'].fill = input_fill
ws_calc['B6'].border = thin_border

ws_calc['A7'] = "Monthly Rate (no commitment) (£)"
ws_calc['B7'] = 350
ws_calc['B7'].number_format = '£#,##0'
ws_calc['B7'].fill = input_fill
ws_calc['B7'].border = thin_border

# Revenue by Number of Firms
ws_calc['A10'] = "REVENUE BY NUMBER OF FIRMS"
ws_calc['A10'].font = subheader_font

calc_headers = ['# Firms', '£250/mo (2yr)', '£300/mo (2yr)', '£250/mo (3yr)', '£300/mo (3yr)', 'Difference', 'Avg MRR']
for col, header in enumerate(calc_headers, start=1):
    cell = ws_calc.cell(row=12, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

firm_counts = [1, 2, 3, 4, 5, 10, 15, 20, 25, 50, 75, 100, 150, 200, 250, 300]
for row_idx, firms in enumerate(firm_counts, start=13):
    ws_calc.cell(row=row_idx, column=1, value=firms).border = thin_border
    ws_calc.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

    cell = ws_calc.cell(row=row_idx, column=2, value=f'=A{row_idx}*$B$5*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    cell = ws_calc.cell(row=row_idx, column=3, value=f'=A{row_idx}*$B$6*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    cell = ws_calc.cell(row=row_idx, column=4, value=f'=A{row_idx}*$B$5*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    cell = ws_calc.cell(row=row_idx, column=5, value=f'=A{row_idx}*$B$6*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    cell = ws_calc.cell(row=row_idx, column=6, value=f'=E{row_idx}-B{row_idx}')
    cell.number_format = '£#,##0'
    cell.border = thin_border
    cell.fill = highlight_fill

    cell = ws_calc.cell(row=row_idx, column=7, value=f'=A{row_idx}*($B$5+$B$6)/2')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    if row_idx % 2 == 0:
        for c in range(1, 6):
            if not ws_calc.cell(row=row_idx, column=c).fill.start_color.rgb or ws_calc.cell(row=row_idx, column=c).fill.start_color.rgb == '00000000':
                ws_calc.cell(row=row_idx, column=c).fill = alt_fill

# Data for Revenue Chart (select key milestones)
ws_calc['I10'] = "Revenue Milestones (for chart)"
ws_calc['I10'].font = section_font

chart_data_headers = ['Firms', '2yr @ £250', '3yr @ £300']
for col, header in enumerate(chart_data_headers, start=9):
    cell = ws_calc.cell(row=11, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Key milestones: 10, 25, 50, 100 firms
milestones = [(10, 17), (25, 20), (50, 21), (100, 24)]  # (firms, row in main table)
for i, (firms, src_row) in enumerate(milestones):
    row = 12 + i
    ws_calc.cell(row=row, column=9, value=firms).border = thin_border
    ws_calc.cell(row=row, column=10, value=f'=B{src_row}').border = thin_border
    ws_calc.cell(row=row, column=10).number_format = '£#,##0'
    ws_calc.cell(row=row, column=11, value=f'=E{src_row}').border = thin_border
    ws_calc.cell(row=row, column=11).number_format = '£#,##0'

# Add line chart for revenue milestones
chart_rev = LineChart()
chart_rev.title = "Total Contract Value by # of Firms"
chart_rev.style = 10
chart_rev.y_axis.title = "Total Contract Value (£)"
chart_rev.x_axis.title = "Number of Firms"
chart_rev.y_axis.numFmt = '£#,##0'

data_rev = Reference(ws_calc, min_col=10, min_row=11, max_col=11, max_row=15)
cats_rev = Reference(ws_calc, min_col=9, min_row=12, max_row=15)
chart_rev.add_data(data_rev, titles_from_data=True)
chart_rev.set_categories(cats_rev)
chart_rev.width = 12
chart_rev.height = 9

ws_calc.add_chart(chart_rev, "I17")

# Column widths
for col in range(1, 12):
    ws_calc.column_dimensions[get_column_letter(col)].width = 15

# ============================================
# SHEET 4: GROWTH PROJECTIONS (with Chart)
# ============================================
ws_growth = wb.create_sheet("Growth Projections")

ws_growth['A1'] = "GROWTH PROJECTIONS"
ws_growth['A1'].font = title_font
ws_growth.merge_cells('A1:F1')

# Input parameters
ws_growth['A3'] = "SCENARIO INPUTS (Edit yellow cells)"
ws_growth['A3'].font = subheader_font

ws_growth['A5'] = "Monthly Rate (£)"
ws_growth['B5'] = 250
ws_growth['B5'].number_format = '£#,##0'
ws_growth['B5'].fill = input_fill
ws_growth['B5'].border = thin_border

ws_growth['A6'] = "Annual Churn Rate (%)"
ws_growth['B6'] = 0.05
ws_growth['B6'].number_format = '0%'
ws_growth['B6'].fill = input_fill
ws_growth['B6'].border = thin_border

growth_headers = ['Year', 'New Firms', 'Churn', 'Total Firms', 'MRR', 'ARR']

# Conservative Growth
ws_growth['A9'] = "CONSERVATIVE (10 new firms/year)"
ws_growth['A9'].font = section_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=10, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

conservative_new = [10, 10, 15, 20, 25]
for i, new_firms in enumerate(conservative_new):
    row = 11 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Moderate Growth
ws_growth['A18'] = "MODERATE (25 new firms/year)"
ws_growth['A18'].font = section_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=19, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

moderate_new = [25, 35, 45, 60, 75]
for i, new_firms in enumerate(moderate_new):
    row = 20 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Aggressive Growth
ws_growth['A27'] = "AGGRESSIVE (50 new firms/year)"
ws_growth['A27'].font = section_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=28, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

aggressive_new = [50, 70, 100, 130, 150]
for i, new_firms in enumerate(aggressive_new):
    row = 29 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Chart data table
ws_growth['H3'] = "ARR Comparison (for chart)"
ws_growth['H3'].font = section_font

chart_headers = ['Year', 'Conservative', 'Moderate', 'Aggressive']
for col, header in enumerate(chart_headers, start=8):
    cell = ws_growth.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i in range(5):
    row = 5 + i
    ws_growth.cell(row=row, column=8, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=9, value=f'=F{11+i}').border = thin_border
    ws_growth.cell(row=row, column=9).number_format = '£#,##0'
    ws_growth.cell(row=row, column=10, value=f'=F{20+i}').border = thin_border
    ws_growth.cell(row=row, column=10).number_format = '£#,##0'
    ws_growth.cell(row=row, column=11, value=f'=F{29+i}').border = thin_border
    ws_growth.cell(row=row, column=11).number_format = '£#,##0'

# Line chart for ARR
chart3 = LineChart()
chart3.title = "ARR Growth Projections"
chart3.style = 10
chart3.y_axis.title = "Annual Recurring Revenue (£)"
chart3.x_axis.title = "Year"
chart3.y_axis.numFmt = '£#,##0'

data3 = Reference(ws_growth, min_col=9, min_row=4, max_col=11, max_row=9)
cats3 = Reference(ws_growth, min_col=8, min_row=5, max_row=9)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 14
chart3.height = 10

ws_growth.add_chart(chart3, "H12")

# Column widths
for col in range(1, 12):
    ws_growth.column_dimensions[get_column_letter(col)].width = 14

# ============================================
# SHEET 5: ROI CALCULATOR (with improved chart)
# ============================================
ws_roi = wb.create_sheet("ROI Calculator")

ws_roi['A1'] = "CLIENT ROI CALCULATOR"
ws_roi['A1'].font = title_font
ws_roi.merge_cells('A1:D1')

ws_roi['A2'] = "Calculate the ROI for your prospective clients"
ws_roi['A2'].font = small_font

ws_roi['A4'] = "CURRENT TOOL COSTS (Edit yellow cells)"
ws_roi['A4'].font = subheader_font

inputs = [
    ['Tool', 'Monthly Cost', 'Notes'],
    ['CRM (generic)', 50, 'Salesforce/HubSpot equivalent'],
    ['Risk Profiling Tool', 50, 'Dynamic Planner element'],
    ['Cash Flow (Voyant)', 175, 'Verified Dec 2025'],
    ['Monte Carlo (Timeline)', 162, 'Verified Dec 2025 (£135+VAT)'],
    ['Document Generation', 50, 'Word templates/Templafy'],
    ['E-Signatures', 23, 'DocuSign/Adobe Sign'],
    ['Compliance Tracking', 50, 'Manual/specialist tool'],
]

for row_idx, row_data in enumerate(inputs, start=6):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 6:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 2:
                cell.fill = input_fill
                cell.number_format = '£#,##0'
        cell.border = thin_border

# Total current cost
ws_roi['A14'] = "TOTAL CURRENT MONTHLY COST"
ws_roi['A14'].font = Font(name='Calibri', size=10, bold=True)
ws_roi['B14'] = '=SUM(B7:B13)'
ws_roi['B14'].number_format = '£#,##0'
ws_roi['B14'].font = Font(name='Calibri', size=10, bold=True)
ws_roi['A14'].fill = blue_fill
ws_roi['B14'].fill = blue_fill
ws_roi['A14'].border = thin_border
ws_roi['B14'].border = thin_border

# Time savings section
ws_roi['A17'] = "TIME SAVINGS (Edit yellow cells)"
ws_roi['A17'].font = subheader_font

time_inputs = [
    ['Metric', 'Value', 'Notes'],
    ['Hours per client onboarding (current)', 15, 'Manual process'],
    ['Time saved with Plannetic (%)', 0.6, '60% automation'],
    ['Your hourly rate (£)', 100, 'Advisor charge-out rate'],
    ['New clients per month', 4, 'Average new clients'],
]

for row_idx, row_data in enumerate(time_inputs, start=19):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 19:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 2:
                cell.fill = input_fill
                if row_idx == 21:
                    cell.number_format = '0%'
                elif row_idx == 22:
                    cell.number_format = '£#,##0'
        cell.border = thin_border

# Plannetic cost
ws_roi['A26'] = "PLANNETIC COST"
ws_roi['A26'].font = subheader_font

ws_roi['A28'] = "Plannetic Monthly Cost"
ws_roi['B28'] = 250
ws_roi['B28'].number_format = '£#,##0'
ws_roi['B28'].fill = input_fill
ws_roi['B28'].border = thin_border
ws_roi['A28'].border = thin_border

# Results section
ws_roi['A31'] = "ROI ANALYSIS"
ws_roi['A31'].font = subheader_font

results = [
    ['Metric', 'Monthly', 'Annual'],
    ['Software Savings', '=B14-B28', '=B33*12'],
    ['Hours Saved per Client', '=B20*B21', ''],
    ['Total Hours Saved (all clients)', '=B34*B23', '=B35*12'],
    ['Value of Time Saved', '=B35*B22', '=B36*12'],
    ['', '', ''],
    ['TOTAL BENEFIT', '=B33+B36', '=B38*12'],
    ['Plannetic Cost', '=B28', '=B28*12'],
    ['NET BENEFIT', '=B38-B39', '=B40*12'],
    ['ROI %', '=(B40/B39)*100', ''],
]

for row_idx, row_data in enumerate(results, start=33):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 33:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx in [2, 3] and value:
                if row_idx == 43:
                    cell.number_format = '0"%"'
                elif row_idx in [34, 35, 36]:
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '£#,##0'
        cell.border = thin_border

        if row_idx in [38, 40, 41, 43]:
            cell.fill = highlight_fill
            cell.font = Font(name='Calibri', size=10, bold=True)

# Chart data for ROI visualization
ws_roi['E31'] = "Savings Breakdown"
ws_roi['E31'].font = section_font

ws_roi['E33'] = "Category"
ws_roi['F33'] = "Monthly"
ws_roi['G33'] = "Annual"
ws_roi['E33'].font = header_font
ws_roi['F33'].font = header_font
ws_roi['G33'].font = header_font
ws_roi['E33'].fill = header_fill
ws_roi['F33'].fill = header_fill
ws_roi['G33'].fill = header_fill
ws_roi['E33'].border = thin_border
ws_roi['F33'].border = thin_border
ws_roi['G33'].border = thin_border

ws_roi['E34'] = "Software Savings"
ws_roi['F34'] = '=B33'
ws_roi['G34'] = '=C33'
ws_roi['E35'] = "Time Value Saved"
ws_roi['F35'] = '=B36'
ws_roi['G35'] = '=C36'
ws_roi['E36'] = "Total Benefit"
ws_roi['F36'] = '=B38'
ws_roi['G36'] = '=C38'
ws_roi['E37'] = "Plannetic Cost"
ws_roi['F37'] = '=B39'
ws_roi['G37'] = '=C39'
ws_roi['E38'] = "Net Benefit"
ws_roi['F38'] = '=B40'
ws_roi['G38'] = '=C40'

for row in range(34, 39):
    for col in range(5, 8):
        cell = ws_roi.cell(row=row, column=col)
        cell.border = thin_border
        if col in [6, 7]:
            cell.number_format = '£#,##0'
        if row == 38:
            cell.fill = highlight_fill
            cell.font = Font(name='Calibri', size=10, bold=True)

# Stacked bar chart showing benefit vs cost
chart_roi = BarChart()
chart_roi.type = "col"
chart_roi.style = 10
chart_roi.title = "Monthly Value Analysis"
chart_roi.y_axis.title = "£ per month"
chart_roi.y_axis.numFmt = '£#,##0'

# Use software savings, time savings, and cost
data_roi = Reference(ws_roi, min_col=6, min_row=33, max_row=37)
cats_roi = Reference(ws_roi, min_col=5, min_row=34, max_row=37)
chart_roi.add_data(data_roi, titles_from_data=True)
chart_roi.set_categories(cats_roi)
chart_roi.width = 11
chart_roi.height = 9
chart_roi.legend = None

ws_roi.add_chart(chart_roi, "E41")

# Column widths
ws_roi.column_dimensions['A'].width = 32
ws_roi.column_dimensions['B'].width = 14
ws_roi.column_dimensions['C'].width = 14
ws_roi.column_dimensions['D'].width = 8
ws_roi.column_dimensions['E'].width = 18
ws_roi.column_dimensions['F'].width = 12
ws_roi.column_dimensions['G'].width = 12

# ============================================
# SHEET 6: TIER COMPARISON
# ============================================
ws_tiers = wb.create_sheet("Tier Comparison")

ws_tiers['A1'] = "PLANNETIC PRICING TIERS"
ws_tiers['A1'].font = title_font
ws_tiers.merge_cells('A1:D1')

tier_headers = ['Feature', 'Standard £250/mo', 'Professional £300/mo', 'Enterprise (Custom)']
for col, header in enumerate(tier_headers, start=1):
    cell = ws_tiers.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

features = [
    ['CORE FEATURES', '', '', ''],
    ['Client Management Hub', '✓', '✓', '✓'],
    ['All 6 Assessment Types', '✓', '✓', '✓'],
    ['Unlimited Clients', '✓', '✓', '✓'],
    ['Document Generation', '✓', '✓', '✓'],
    ['E-Signatures', 'Fair Use', 'Unlimited', 'Unlimited'],
    ['FCA Compliance Registers', '✓', '✓', '✓'],
    ['Consumer Duty Workflows', '✓', '✓', '✓'],
    ['', '', '', ''],
    ['PREMIUM FEATURES', '', '', ''],
    ['AI Enhancement', '—', '✓', '✓'],
    ['Priority Support (4hr SLA)', '—', '✓', '✓'],
    ['Phone Support', '—', '✓', '✓'],
    ['White-label Client Portal', '—', '✓', '✓'],
    ['Advanced Analytics', '—', '✓', '✓'],
    ['API Access', '—', '—', '✓'],
    ['Custom Integrations', '—', '—', '✓'],
    ['Dedicated Account Manager', '—', '—', '✓'],
    ['', '', '', ''],
    ['ONBOARDING', '', '', ''],
    ['Data Migration', 'Self-serve', 'Assisted', 'Full Service'],
    ['Training Sessions', '2 calls', '4 calls', 'Unlimited'],
    ['', '', '', ''],
    ['CONTRACT', '', '', ''],
    ['Minimum Commitment', '2 years', '2 years', '3 years'],
    ['2-Year TCV', '=250*24', '=300*24', 'Custom'],
    ['3-Year TCV', '=250*36', '=300*36', 'Custom'],
]

for row_idx, row_data in enumerate(features, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_tiers.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

        if value in ['CORE FEATURES', 'PREMIUM FEATURES', 'ONBOARDING', 'CONTRACT']:
            cell.font = section_font
            cell.fill = blue_fill
        elif value == '✓':
            cell.font = Font(name='Calibri', size=10, color='059669', bold=True)
        elif value == '—':
            cell.font = Font(name='Calibri', size=10, color='9CA3AF')
        else:
            cell.font = normal_font

        if row_idx in [29, 30] and col_idx in [2, 3]:
            cell.number_format = '£#,##0'

# Column widths
ws_tiers.column_dimensions['A'].width = 28
ws_tiers.column_dimensions['B'].width = 20
ws_tiers.column_dimensions['C'].width = 20
ws_tiers.column_dimensions['D'].width = 20

# Save workbook
output_path = '/Users/adeomosanya/Downloads/Plannetic-Pricing-Analysis-v3.xlsx'
wb.save(output_path)
print(f"✅ Excel file created: {output_path}")
print("\nCharts included:")
print("1. Summary - Cost comparison bar chart")
print("2. Competitor Pricing - Software costs bar + Tool stack pie")
print("3. Revenue Calculator - TCV line chart by # of firms")
print("4. Growth Projections - ARR line chart (3 scenarios)")
print("5. ROI Calculator - Monthly value analysis bar chart")
print("6. Tier Comparison - Feature matrix (no chart needed)")
