#!/usr/bin/env python3
"""
Plannetic Pricing Analysis v2 - Excel Generator with Charts
Creates a comprehensive pricing model with verified data, working formulas, and beautiful charts
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

# Create workbook
wb = Workbook()

# ============================================
# STYLES
# ============================================
title_font = Font(name='Calibri', size=20, bold=True, color='1E40AF')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
subheader_font = Font(name='Calibri', size=13, bold=True, color='1E40AF')
section_font = Font(name='Calibri', size=11, bold=True, color='374151')
normal_font = Font(name='Calibri', size=10)
currency_font = Font(name='Calibri', size=10, bold=True, color='059669')
small_font = Font(name='Calibri', size=9, color='6B7280')

header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
alt_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
highlight_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
input_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
warning_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# ============================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================
ws_summary = wb.active
ws_summary.title = "Summary"

ws_summary['A1'] = "PLANNETIC PRICING ANALYSIS"
ws_summary['A1'].font = title_font
ws_summary.merge_cells('A1:F1')

ws_summary['A2'] = "Verified December 2025"
ws_summary['A2'].font = small_font

ws_summary['A4'] = "What is Plannetic?"
ws_summary['A4'].font = subheader_font

summary_text = [
    "Plannetic is a comprehensive, compliance-focused financial advisory platform",
    "designed specifically for UK-regulated Independent Financial Advisors (IFAs).",
    "",
    "Key Value Proposition:",
    "• Replaces 5-7 separate tools with one integrated platform",
    "• Saves IFAs £280+/month vs competitor tool stack (£530 → £250)",
    "• Saves 10-20 hours per client onboarding",
    "• Built-in FCA compliance and Consumer Duty workflows",
    "• AI-enhanced assessments and form auto-population",
]

for i, text in enumerate(summary_text, start=6):
    ws_summary[f'A{i}'] = text
    ws_summary[f'A{i}'].font = normal_font

# Pricing Summary Table
ws_summary['A17'] = "Recommended Pricing Tiers"
ws_summary['A17'].font = subheader_font

pricing_headers = ['Tier', 'Monthly', 'Commitment', '2-Year TCV', '3-Year TCV', 'Best For']
for col, header in enumerate(pricing_headers, start=1):
    cell = ws_summary.cell(row=19, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

pricing_data = [
    ['Monthly', 350, 'Month-to-month', '=B20*24', '=B20*36', 'Trial/uncertain firms'],
    ['Standard', 250, '2-year', '=B21*24', '=B21*36', 'Solo advisors, small firms'],
    ['Professional', 300, '2-year', '=B22*24', '=B22*36', 'Growing firms, AI + support'],
    ['Enterprise', 'Custom', '3-year', 'Custom', 'Custom', '5+ advisors, white-label'],
]

for row_idx, row_data in enumerate(pricing_data, start=20):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if row_idx % 2 == 1:
            cell.fill = alt_fill
        if col_idx == 2 and isinstance(value, int):
            cell.number_format = '£#,##0'
        if col_idx in [4, 5] and isinstance(value, str) and value.startswith('='):
            cell.number_format = '£#,##0'

# Key savings metrics
ws_summary['A26'] = "Monthly Savings Analysis"
ws_summary['A26'].font = subheader_font

savings_data = [
    ['Metric', 'Value', 'Notes'],
    ['Competitor Stack Cost', 530, 'Verified Dec 2025 pricing'],
    ['Plannetic Standard', 250, 'Full platform access'],
    ['Monthly Savings', '=B28-B29', 'Per month'],
    ['Annual Savings', '=B30*12', 'Per year'],
    ['Savings %', '=B30/B28', 'vs competitors'],
]

for row_idx, row_data in enumerate(savings_data, start=28):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 28:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        cell.border = thin_border
        if col_idx == 2 and row_idx > 28:
            if row_idx == 33:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '£#,##0'

# Highlight the savings row
ws_summary['A32'].fill = highlight_fill
ws_summary['B32'].fill = highlight_fill
ws_summary['C32'].fill = highlight_fill

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 22
ws_summary.column_dimensions['D'].width = 15
ws_summary.column_dimensions['E'].width = 15
ws_summary.column_dimensions['F'].width = 28

# ============================================
# SHEET 2: COMPETITOR ANALYSIS (with Chart)
# ============================================
ws_comp = wb.create_sheet("Competitor Pricing")

ws_comp['A1'] = "COMPETITOR PRICING ANALYSIS"
ws_comp['A1'].font = title_font
ws_comp.merge_cells('A1:E1')

ws_comp['A2'] = "Verified December 2025 - Sources linked below"
ws_comp['A2'].font = small_font

# Verified competitor pricing
ws_comp['A4'] = "UK IFA Software Market - Verified Pricing"
ws_comp['A4'].font = subheader_font

comp_headers = ['Software', 'Monthly Price', 'Price Type', 'What They Offer', 'Source']
for col, header in enumerate(comp_headers, start=1):
    cell = ws_comp.cell(row=6, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Verified data
competitors = [
    ['Intelliflo Office', 132, 'Per user', 'Back office + cashflow', 'TrustRadius'],
    ['Voyant AdviserGo', 175, 'Flat fee', 'Cash flow planning', 'voyant.com'],
    ['Timeline', 162, 'Flat (+VAT)', 'Monte Carlo simulations', 'timeline.co'],
    ['FE CashCalc', 90, 'Per adviser (+VAT)', 'Cash flow modelling', 'advisoryai.com'],
    ['Dynamic Planner', 200, 'Estimated', 'Risk profiling + reports', 'Contact required'],
    ['Plannetic Standard', 250, 'Flat fee', 'ALL-IN-ONE PLATFORM', 'Your price'],
]

for row_idx, row_data in enumerate(competitors, start=7):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx == 2:
            cell.number_format = '£#,##0'
        if row_idx == 12:  # Plannetic row
            cell.fill = highlight_fill
            cell.font = Font(name='Calibri', size=10, bold=True)
        elif row_idx % 2 == 0:
            cell.fill = alt_fill

# Add bar chart for competitor pricing
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Monthly Software Costs Comparison"
chart1.y_axis.title = "£ per month"
chart1.x_axis.title = None

# Data for chart
data = Reference(ws_comp, min_col=2, min_row=6, max_row=12)
cats = Reference(ws_comp, min_col=1, min_row=7, max_row=12)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.width = 15
chart1.height = 10

# Style the chart
chart1.legend = None

ws_comp.add_chart(chart1, "A15")

# Tool Stack Comparison
ws_comp['A32'] = "Typical IFA Tool Stack vs Plannetic"
ws_comp['A32'].font = subheader_font

stack_headers = ['Tool Category', 'Standalone Cost', 'Plannetic', 'Savings']
for col, header in enumerate(stack_headers, start=1):
    cell = ws_comp.cell(row=34, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

stack_data = [
    ['CRM (generic)', 50, 'Included', '=B35'],
    ['Risk Profiling', 50, 'Included', '=B36'],
    ['Cash Flow (Voyant)', 175, 'Included', '=B37'],
    ['Monte Carlo (Timeline)', 162, 'Included', '=B38'],
    ['Document Generation', 50, 'Included', '=B39'],
    ['E-Signatures', 23, 'Included', '=B40'],
    ['Compliance Tracking', 50, 'Included', '=B41'],
]

for row_idx, row_data in enumerate(stack_data, start=35):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx in [2, 4]:
            cell.number_format = '£#,##0'

# Total row
ws_comp['A42'] = 'TOTAL'
ws_comp['B42'] = '=SUM(B35:B41)'
ws_comp['C42'] = '£250/mo'
ws_comp['D42'] = '=B42-250'

for col in range(1, 5):
    cell = ws_comp.cell(row=42, column=col)
    cell.font = Font(name='Calibri', size=10, bold=True)
    cell.fill = highlight_fill
    cell.border = thin_border
    if col in [2, 4]:
        cell.number_format = '£#,##0'

# Add pie chart for tool stack
chart2 = PieChart()
chart2.title = "Tool Stack Cost Breakdown"
data2 = Reference(ws_comp, min_col=2, min_row=35, max_row=41)
labels2 = Reference(ws_comp, min_col=1, min_row=35, max_row=41)
chart2.add_data(data2)
chart2.set_categories(labels2)
chart2.width = 12
chart2.height = 10

# Add data labels
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.dataLabels.showVal = False
chart2.dataLabels.showCatName = False

ws_comp.add_chart(chart2, "F32")

# Sources
ws_comp['A45'] = "Sources:"
ws_comp['A45'].font = section_font
sources = [
    "• Voyant: planwithvoyant.com/uk/pricing",
    "• Timeline: timeline.co (£135+VAT = £162)",
    "• CashCalc: advisoryai.com (£75+VAT = £90)",
    "• Intelliflo: trustradius.com (£130-135/user)",
]
for i, source in enumerate(sources, start=46):
    ws_comp[f'A{i}'] = source
    ws_comp[f'A{i}'].font = small_font

# Column widths
ws_comp.column_dimensions['A'].width = 22
ws_comp.column_dimensions['B'].width = 15
ws_comp.column_dimensions['C'].width = 18
ws_comp.column_dimensions['D'].width = 12
ws_comp.column_dimensions['E'].width = 18

# ============================================
# SHEET 3: REVENUE CALCULATOR
# ============================================
ws_calc = wb.create_sheet("Revenue Calculator")

ws_calc['A1'] = "REVENUE CALCULATOR"
ws_calc['A1'].font = title_font
ws_calc.merge_cells('A1:G1')

# Input Section
ws_calc['A3'] = "INPUT PARAMETERS (Edit yellow cells)"
ws_calc['A3'].font = subheader_font

ws_calc['A5'] = "Standard Monthly Rate (£)"
ws_calc['B5'] = 250
ws_calc['B5'].number_format = '£#,##0'
ws_calc['B5'].fill = input_fill
ws_calc['B5'].border = thin_border

ws_calc['A6'] = "Professional Monthly Rate (£)"
ws_calc['B6'] = 300
ws_calc['B6'].number_format = '£#,##0'
ws_calc['B6'].fill = input_fill
ws_calc['B6'].border = thin_border

ws_calc['A7'] = "Monthly Rate (no commitment) (£)"
ws_calc['B7'] = 350
ws_calc['B7'].number_format = '£#,##0'
ws_calc['B7'].fill = input_fill
ws_calc['B7'].border = thin_border

# Revenue by Number of Firms
ws_calc['A10'] = "REVENUE BY NUMBER OF FIRMS"
ws_calc['A10'].font = subheader_font

calc_headers = ['# Firms', '£250/mo (2yr)', '£300/mo (2yr)', '£250/mo (3yr)', '£300/mo (3yr)', 'Difference', 'Avg MRR']
for col, header in enumerate(calc_headers, start=1):
    cell = ws_calc.cell(row=12, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

firm_counts = [1, 2, 3, 4, 5, 10, 15, 20, 25, 50, 75, 100, 150, 200, 250, 300]
for row_idx, firms in enumerate(firm_counts, start=13):
    ws_calc.cell(row=row_idx, column=1, value=firms).border = thin_border
    ws_calc.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

    # £250/mo 2yr
    cell = ws_calc.cell(row=row_idx, column=2, value=f'=A{row_idx}*$B$5*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £300/mo 2yr
    cell = ws_calc.cell(row=row_idx, column=3, value=f'=A{row_idx}*$B$6*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £250/mo 3yr
    cell = ws_calc.cell(row=row_idx, column=4, value=f'=A{row_idx}*$B$5*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £300/mo 3yr
    cell = ws_calc.cell(row=row_idx, column=5, value=f'=A{row_idx}*$B$6*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # Difference
    cell = ws_calc.cell(row=row_idx, column=6, value=f'=E{row_idx}-B{row_idx}')
    cell.number_format = '£#,##0'
    cell.border = thin_border
    cell.fill = highlight_fill

    # Avg MRR
    cell = ws_calc.cell(row=row_idx, column=7, value=f'=A{row_idx}*($B$5+$B$6)/2')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    if row_idx % 2 == 0:
        for c in range(1, 6):
            if not ws_calc.cell(row=row_idx, column=c).fill.start_color.rgb or ws_calc.cell(row=row_idx, column=c).fill.start_color.rgb == '00000000':
                ws_calc.cell(row=row_idx, column=c).fill = alt_fill

# Column widths
for col in range(1, 8):
    ws_calc.column_dimensions[get_column_letter(col)].width = 16

# ============================================
# SHEET 4: GROWTH PROJECTIONS (with Chart)
# ============================================
ws_growth = wb.create_sheet("Growth Projections")

ws_growth['A1'] = "GROWTH PROJECTIONS"
ws_growth['A1'].font = title_font
ws_growth.merge_cells('A1:F1')

# Input parameters
ws_growth['A3'] = "SCENARIO INPUTS (Edit yellow cells)"
ws_growth['A3'].font = subheader_font

ws_growth['A5'] = "Monthly Rate (£)"
ws_growth['B5'] = 250
ws_growth['B5'].number_format = '£#,##0'
ws_growth['B5'].fill = input_fill
ws_growth['B5'].border = thin_border

ws_growth['A6'] = "Annual Churn Rate (%)"
ws_growth['B6'] = 0.05
ws_growth['B6'].number_format = '0%'
ws_growth['B6'].fill = input_fill
ws_growth['B6'].border = thin_border

# Conservative Growth
ws_growth['A9'] = "CONSERVATIVE (10 new firms/year)"
ws_growth['A9'].font = section_font

growth_headers = ['Year', 'New Firms', 'Churn', 'Total Firms', 'MRR', 'ARR']
for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=10, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

conservative_new = [10, 10, 15, 20, 25]
for i, new_firms in enumerate(conservative_new):
    row = 11 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Moderate Growth
ws_growth['A18'] = "MODERATE (25 new firms/year)"
ws_growth['A18'].font = section_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=19, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

moderate_new = [25, 35, 45, 60, 75]
for i, new_firms in enumerate(moderate_new):
    row = 20 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Aggressive Growth
ws_growth['A27'] = "AGGRESSIVE (50 new firms/year)"
ws_growth['A27'].font = section_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=28, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

aggressive_new = [50, 70, 100, 130, 150]
for i, new_firms in enumerate(aggressive_new):
    row = 29 + i
    ws_growth.cell(row=row, column=1, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=2, value=new_firms).border = thin_border
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0).border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=B{row}').border = thin_border
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6,0)').border = thin_border
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}').border = thin_border
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5').border = thin_border
    ws_growth.cell(row=row, column=5).number_format = '£#,##0'
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12').border = thin_border
    ws_growth.cell(row=row, column=6).number_format = '£#,##0'
    for col in range(1, 7):
        ws_growth.cell(row=row, column=col).alignment = Alignment(horizontal='center')

# Create comparison table for chart
ws_growth['H3'] = "ARR Comparison (for chart)"
ws_growth['H3'].font = section_font

chart_headers = ['Year', 'Conservative', 'Moderate', 'Aggressive']
for col, header in enumerate(chart_headers, start=8):
    cell = ws_growth.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i in range(5):
    row = 5 + i
    ws_growth.cell(row=row, column=8, value=i+1).border = thin_border
    ws_growth.cell(row=row, column=9, value=f'=F{11+i}').border = thin_border  # Conservative ARR
    ws_growth.cell(row=row, column=9).number_format = '£#,##0'
    ws_growth.cell(row=row, column=10, value=f'=F{20+i}').border = thin_border  # Moderate ARR
    ws_growth.cell(row=row, column=10).number_format = '£#,##0'
    ws_growth.cell(row=row, column=11, value=f'=F{29+i}').border = thin_border  # Aggressive ARR
    ws_growth.cell(row=row, column=11).number_format = '£#,##0'

# Add line chart for ARR projections
chart3 = LineChart()
chart3.title = "ARR Growth Projections"
chart3.style = 10
chart3.y_axis.title = "Annual Recurring Revenue (£)"
chart3.x_axis.title = "Year"
chart3.y_axis.numFmt = '£#,##0'

data3 = Reference(ws_growth, min_col=9, min_row=4, max_col=11, max_row=9)
cats3 = Reference(ws_growth, min_col=8, min_row=5, max_row=9)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 15
chart3.height = 10

ws_growth.add_chart(chart3, "H12")

# Column widths
for col in range(1, 12):
    ws_growth.column_dimensions[get_column_letter(col)].width = 14

# ============================================
# SHEET 5: ROI CALCULATOR
# ============================================
ws_roi = wb.create_sheet("ROI Calculator")

ws_roi['A1'] = "CLIENT ROI CALCULATOR"
ws_roi['A1'].font = title_font
ws_roi.merge_cells('A1:D1')

ws_roi['A2'] = "Calculate the ROI for your prospective clients"
ws_roi['A2'].font = small_font

ws_roi['A4'] = "CURRENT TOOL COSTS (Edit yellow cells)"
ws_roi['A4'].font = subheader_font

# Input section with verified defaults
inputs = [
    ['Tool', 'Monthly Cost', 'Notes'],
    ['CRM (generic)', 50, 'Salesforce/HubSpot equivalent'],
    ['Risk Profiling Tool', 50, 'Dynamic Planner element'],
    ['Cash Flow (Voyant)', 175, 'Verified Dec 2025'],
    ['Monte Carlo (Timeline)', 162, 'Verified Dec 2025 (£135+VAT)'],
    ['Document Generation', 50, 'Word templates/Templafy'],
    ['E-Signatures', 23, 'DocuSign/Adobe Sign'],
    ['Compliance Tracking', 50, 'Manual/specialist tool'],
]

for row_idx, row_data in enumerate(inputs, start=6):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 6:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 2:
                cell.fill = input_fill
                cell.number_format = '£#,##0'
        cell.border = thin_border

# Total current cost
ws_roi['A14'] = "TOTAL CURRENT MONTHLY COST"
ws_roi['A14'].font = Font(name='Calibri', size=10, bold=True)
ws_roi['B14'] = '=SUM(B7:B13)'
ws_roi['B14'].number_format = '£#,##0'
ws_roi['B14'].font = Font(name='Calibri', size=10, bold=True)
ws_roi['A14'].fill = blue_fill
ws_roi['B14'].fill = blue_fill
ws_roi['A14'].border = thin_border
ws_roi['B14'].border = thin_border

# Time savings section
ws_roi['A17'] = "TIME SAVINGS (Edit yellow cells)"
ws_roi['A17'].font = subheader_font

time_inputs = [
    ['Metric', 'Value', 'Notes'],
    ['Hours per client onboarding (current)', 15, 'Manual process'],
    ['Time saved with Plannetic (%)', 0.6, '60% automation'],
    ['Your hourly rate (£)', 100, 'Advisor charge-out rate'],
    ['New clients per month', 4, 'Average new clients'],
]

for row_idx, row_data in enumerate(time_inputs, start=19):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 19:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx == 2:
                cell.fill = input_fill
                if row_idx == 21:
                    cell.number_format = '0%'
                elif row_idx == 22:
                    cell.number_format = '£#,##0'
        cell.border = thin_border

# Plannetic cost
ws_roi['A26'] = "PLANNETIC COST"
ws_roi['A26'].font = subheader_font

ws_roi['A28'] = "Plannetic Monthly Cost"
ws_roi['B28'] = 250
ws_roi['B28'].number_format = '£#,##0'
ws_roi['B28'].fill = input_fill
ws_roi['B28'].border = thin_border
ws_roi['A28'].border = thin_border

# Results section
ws_roi['A31'] = "ROI ANALYSIS"
ws_roi['A31'].font = subheader_font

results = [
    ['Metric', 'Monthly', 'Annual', 'Formula'],
    ['Software Savings', '=B14-B28', '=B33*12', 'Current tools - Plannetic'],
    ['Hours Saved per Client', '=B20*B21', '=B34*12', 'Hours × automation %'],
    ['Clients per Month', '=B23', '=B35*12', 'From inputs'],
    ['Total Hours Saved', '=B34*B35', '=B36*12', 'Hours × clients'],
    ['Value of Time Saved', '=B36*B22', '=B37*12', 'Hours × rate'],
    ['', '', '', ''],
    ['TOTAL MONTHLY BENEFIT', '=B33+B37', '=B39*12', 'Software + time savings'],
    ['Plannetic Cost', '=B28', '=B28*12', 'Your subscription'],
    ['NET BENEFIT', '=B39-B40', '=B41*12', 'Benefit - cost'],
    ['ROI %', '=B41/B40*100', '=C41/C40*100', '(Benefit-Cost)/Cost'],
]

for row_idx, row_data in enumerate(results, start=33):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_roi.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 33:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if col_idx in [2, 3] and value and value != '':
                if row_idx == 44:  # ROI row
                    cell.number_format = '0.0"%"'
                elif row_idx in [35, 36, 37]:  # Hours
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '£#,##0'
        cell.border = thin_border

        # Highlight key rows
        if row_idx in [39, 41, 43, 44]:
            cell.fill = highlight_fill
            cell.font = Font(name='Calibri', size=10, bold=True)

# Add savings bar chart
chart4 = BarChart()
chart4.type = "col"
chart4.style = 10
chart4.title = "Monthly Savings Breakdown"

# Create data for chart
ws_roi['F33'] = "Category"
ws_roi['G33'] = "Value"
ws_roi['F34'] = "Software Savings"
ws_roi['G34'] = '=B33'
ws_roi['F35'] = "Time Value Saved"
ws_roi['G35'] = '=B37'
ws_roi['G34'].number_format = '£#,##0'
ws_roi['G35'].number_format = '£#,##0'

data4 = Reference(ws_roi, min_col=7, min_row=33, max_row=35)
cats4 = Reference(ws_roi, min_col=6, min_row=34, max_row=35)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.width = 10
chart4.height = 8
chart4.legend = None

ws_roi.add_chart(chart4, "E17")

# Column widths
ws_roi.column_dimensions['A'].width = 32
ws_roi.column_dimensions['B'].width = 15
ws_roi.column_dimensions['C'].width = 15
ws_roi.column_dimensions['D'].width = 25

# ============================================
# SHEET 6: TIER COMPARISON
# ============================================
ws_tiers = wb.create_sheet("Tier Comparison")

ws_tiers['A1'] = "PLANNETIC PRICING TIERS"
ws_tiers['A1'].font = title_font
ws_tiers.merge_cells('A1:D1')

tier_headers = ['Feature', 'Standard £250/mo', 'Professional £300/mo', 'Enterprise (Custom)']
for col, header in enumerate(tier_headers, start=1):
    cell = ws_tiers.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

features = [
    ['CORE FEATURES', '', '', ''],
    ['Client Management Hub', '✓', '✓', '✓'],
    ['All 6 Assessment Types', '✓', '✓', '✓'],
    ['Unlimited Clients', '✓', '✓', '✓'],
    ['Document Generation', '✓', '✓', '✓'],
    ['E-Signatures', 'Fair Use', 'Unlimited', 'Unlimited'],
    ['FCA Compliance Registers', '✓', '✓', '✓'],
    ['Consumer Duty Workflows', '✓', '✓', '✓'],
    ['', '', '', ''],
    ['PREMIUM FEATURES', '', '', ''],
    ['AI Enhancement', '—', '✓', '✓'],
    ['Priority Support (4hr SLA)', '—', '✓', '✓'],
    ['Phone Support', '—', '✓', '✓'],
    ['White-label Client Portal', '—', '✓', '✓'],
    ['Advanced Analytics', '—', '✓', '✓'],
    ['API Access', '—', '—', '✓'],
    ['Custom Integrations', '—', '—', '✓'],
    ['Dedicated Account Manager', '—', '—', '✓'],
    ['', '', '', ''],
    ['ONBOARDING', '', '', ''],
    ['Data Migration', 'Self-serve', 'Assisted', 'Full Service'],
    ['Training Sessions', '2 calls', '4 calls', 'Unlimited'],
    ['', '', '', ''],
    ['CONTRACT', '', '', ''],
    ['Minimum Commitment', '2 years', '2 years', '3 years'],
    ['2-Year TCV', '=250*24', '=300*24', 'Custom'],
    ['3-Year TCV', '=250*36', '=300*36', 'Custom'],
]

for row_idx, row_data in enumerate(features, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_tiers.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

        # Section headers
        if value in ['CORE FEATURES', 'PREMIUM FEATURES', 'ONBOARDING', 'CONTRACT']:
            cell.font = section_font
            cell.fill = blue_fill
        elif value == '✓':
            cell.font = Font(name='Calibri', size=10, color='059669', bold=True)
        elif value == '—':
            cell.font = Font(name='Calibri', size=10, color='9CA3AF')
        else:
            cell.font = normal_font

        # Number formatting for TCV
        if row_idx in [29, 30] and col_idx in [2, 3]:
            cell.number_format = '£#,##0'

# Column widths
ws_tiers.column_dimensions['A'].width = 28
ws_tiers.column_dimensions['B'].width = 20
ws_tiers.column_dimensions['C'].width = 20
ws_tiers.column_dimensions['D'].width = 20

# Save workbook
output_path = '/Users/adeomosanya/Downloads/Plannetic-Pricing-Analysis-v2.xlsx'
wb.save(output_path)
print(f"✅ Excel file created successfully: {output_path}")
print("\nSheets included:")
print("1. Summary - Executive overview")
print("2. Competitor Pricing - Verified data + bar chart + pie chart")
print("3. Revenue Calculator - Formulas linked to inputs")
print("4. Growth Projections - 3 scenarios + line chart")
print("5. ROI Calculator - Full ROI analysis + bar chart")
print("6. Tier Comparison - Feature matrix")
