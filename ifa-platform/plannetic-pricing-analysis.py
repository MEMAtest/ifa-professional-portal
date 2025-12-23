#!/usr/bin/env python3
"""
Plannetic Pricing Analysis - Excel Generator
Creates a comprehensive pricing model with working formulas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Create workbook
wb = Workbook()

# ============================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================
ws_summary = wb.active
ws_summary.title = "Executive Summary"

# Styles
title_font = Font(name='Arial', size=18, bold=True, color='1E40AF')
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
subheader_font = Font(name='Arial', size=11, bold=True)
normal_font = Font(name='Arial', size=10)
currency_font = Font(name='Arial', size=10, bold=True, color='059669')

header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
alt_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
highlight_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws_summary['A1'] = "PLANNETIC PRICING ANALYSIS"
ws_summary['A1'].font = title_font
ws_summary.merge_cells('A1:F1')

ws_summary['A3'] = "Executive Summary"
ws_summary['A3'].font = subheader_font

summary_text = [
    "Plannetic is a comprehensive, compliance-focused financial advisory platform",
    "designed specifically for UK-regulated Independent Financial Advisors (IFAs).",
    "",
    "Key Value Proposition:",
    "• Replaces 5-7 separate tools with one integrated platform",
    "• Saves IFAs £170+/month vs competitor tool stack (£420 → £250)",
    "• Saves 10-20 hours per client onboarding",
    "• Built-in FCA compliance and Consumer Duty workflows",
]

for i, text in enumerate(summary_text, start=5):
    ws_summary[f'A{i}'] = text
    ws_summary[f'A{i}'].font = normal_font

# Pricing Summary Table
ws_summary['A15'] = "Recommended Pricing Tiers"
ws_summary['A15'].font = subheader_font

pricing_headers = ['Tier', 'Monthly Price', 'Commitment', '2-Year TCV', '3-Year TCV', 'Best For']
for col, header in enumerate(pricing_headers, start=1):
    cell = ws_summary.cell(row=17, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

pricing_data = [
    ['Monthly', 350, 'Month-to-month', '=B18*24', '=B18*36', 'Trial/uncertain firms'],
    ['Standard', 250, '2-year', '=B19*24', '=B19*36', 'Solo advisors, small firms'],
    ['Professional', 300, '2-year', '=B20*24', '=B20*36', 'Growing firms, AI + support'],
    ['Enterprise', 'Custom', '3-year', 'Custom', 'Custom', '5+ advisors, white-label'],
]

for row_idx, row_data in enumerate(pricing_data, start=18):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        if col_idx in [2, 4, 5] and isinstance(value, (int, str)) and (isinstance(value, int) or value.startswith('=')):
            cell.number_format = '£#,##0'

# Key metrics
ws_summary['A24'] = "Key Metrics"
ws_summary['A24'].font = subheader_font

metrics = [
    ['Metric', 'Value'],
    ['Competitor Stack Cost', '£420/mo'],
    ['Plannetic Standard', '£250/mo'],
    ['Monthly Savings', '£170/mo'],
    ['Annual Savings', '£2,040/yr'],
    ['Break-even Clients', '3-6 clients'],
]

for row_idx, row_data in enumerate(metrics, start=26):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 26:
            cell.font = header_font
            cell.fill = header_fill
        else:
            cell.font = normal_font
            if row_idx % 2 == 1:
                cell.fill = alt_fill
        cell.border = thin_border

# Column widths
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 15
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 15
ws_summary.column_dimensions['E'].width = 15
ws_summary.column_dimensions['F'].width = 28

# ============================================
# SHEET 2: REVENUE CALCULATOR
# ============================================
ws_calc = wb.create_sheet("Revenue Calculator")

ws_calc['A1'] = "REVENUE CALCULATOR"
ws_calc['A1'].font = title_font
ws_calc.merge_cells('A1:G1')

# Input Section
ws_calc['A3'] = "INPUT PARAMETERS"
ws_calc['A3'].font = subheader_font

ws_calc['A5'] = "Standard Monthly Rate (£)"
ws_calc['B5'] = 250
ws_calc['B5'].number_format = '£#,##0'
ws_calc['B5'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

ws_calc['A6'] = "Professional Monthly Rate (£)"
ws_calc['B6'] = 300
ws_calc['B6'].number_format = '£#,##0'
ws_calc['B6'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

ws_calc['A7'] = "Monthly Rate (no commitment) (£)"
ws_calc['B7'] = 350
ws_calc['B7'].number_format = '£#,##0'
ws_calc['B7'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

# Revenue by Number of Firms
ws_calc['A10'] = "REVENUE BY NUMBER OF FIRMS"
ws_calc['A10'].font = subheader_font

calc_headers = ['# of Firms', '£250/mo (2yr)', '£300/mo (2yr)', '£250/mo (3yr)', '£300/mo (3yr)', '3yr Difference', 'Avg MRR']
for col, header in enumerate(calc_headers, start=1):
    cell = ws_calc.cell(row=12, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

firm_counts = [1, 2, 3, 4, 5, 10, 15, 20, 25, 50, 75, 100, 150, 200, 250, 300]
for row_idx, firms in enumerate(firm_counts, start=13):
    # Firms count
    cell = ws_calc.cell(row=row_idx, column=1, value=firms)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

    # £250/mo 2yr = firms * 250 * 24
    cell = ws_calc.cell(row=row_idx, column=2, value=f'=A{row_idx}*$B$5*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £300/mo 2yr = firms * 300 * 24
    cell = ws_calc.cell(row=row_idx, column=3, value=f'=A{row_idx}*$B$6*24')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £250/mo 3yr = firms * 250 * 36
    cell = ws_calc.cell(row=row_idx, column=4, value=f'=A{row_idx}*$B$5*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # £300/mo 3yr = firms * 300 * 36
    cell = ws_calc.cell(row=row_idx, column=5, value=f'=A{row_idx}*$B$6*36')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    # Difference (3yr £300 - 2yr £250)
    cell = ws_calc.cell(row=row_idx, column=6, value=f'=E{row_idx}-B{row_idx}')
    cell.number_format = '£#,##0'
    cell.border = thin_border
    cell.fill = highlight_fill

    # Avg MRR (assuming mix of tiers)
    cell = ws_calc.cell(row=row_idx, column=7, value=f'=A{row_idx}*($B$5+$B$6)/2')
    cell.number_format = '£#,##0'
    cell.border = thin_border

    if row_idx % 2 == 0:
        for c in range(1, 7):
            if ws_calc.cell(row=row_idx, column=c).fill.start_color.rgb == '00000000':
                ws_calc.cell(row=row_idx, column=c).fill = alt_fill

# Column widths
for col in range(1, 8):
    ws_calc.column_dimensions[get_column_letter(col)].width = 18

# ============================================
# SHEET 3: GROWTH PROJECTIONS
# ============================================
ws_growth = wb.create_sheet("Growth Projections")

ws_growth['A1'] = "GROWTH PROJECTIONS"
ws_growth['A1'].font = title_font
ws_growth.merge_cells('A1:F1')

# Input parameters
ws_growth['A3'] = "SCENARIO INPUTS"
ws_growth['A3'].font = subheader_font

ws_growth['A5'] = "Monthly Rate (£)"
ws_growth['B5'] = 250
ws_growth['B5'].number_format = '£#,##0'
ws_growth['B5'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

ws_growth['A6'] = "Churn Rate (%)"
ws_growth['B6'] = 5
ws_growth['B6'].number_format = '0%'
ws_growth['B6'].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

# Conservative Growth
ws_growth['A9'] = "CONSERVATIVE GROWTH (10 new firms/year)"
ws_growth['A9'].font = subheader_font

growth_headers = ['Year', 'New Firms', 'Churn', 'Total Firms', 'MRR', 'ARR']
for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=11, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Year 1 Conservative
ws_growth['A12'] = 1
ws_growth['B12'] = 10
ws_growth['C12'] = 0
ws_growth['D12'] = '=B12-C12'
ws_growth['E12'] = '=D12*$B$5'
ws_growth['F12'] = '=E12*12'

# Year 2 Conservative
ws_growth['A13'] = 2
ws_growth['B13'] = 10
ws_growth['C13'] = '=ROUND(D12*$B$6/100,0)'
ws_growth['D13'] = '=D12+B13-C13'
ws_growth['E13'] = '=D13*$B$5'
ws_growth['F13'] = '=E13*12'

# Year 3 Conservative
ws_growth['A14'] = 3
ws_growth['B14'] = 15
ws_growth['C14'] = '=ROUND(D13*$B$6/100,0)'
ws_growth['D14'] = '=D13+B14-C14'
ws_growth['E14'] = '=D14*$B$5'
ws_growth['F14'] = '=E14*12'

# Year 4 Conservative
ws_growth['A15'] = 4
ws_growth['B15'] = 20
ws_growth['C15'] = '=ROUND(D14*$B$6/100,0)'
ws_growth['D15'] = '=D14+B15-C15'
ws_growth['E15'] = '=D15*$B$5'
ws_growth['F15'] = '=E15*12'

# Year 5 Conservative
ws_growth['A16'] = 5
ws_growth['B16'] = 25
ws_growth['C16'] = '=ROUND(D15*$B$6/100,0)'
ws_growth['D16'] = '=D15+B16-C16'
ws_growth['E16'] = '=D16*$B$5'
ws_growth['F16'] = '=E16*12'

for row in range(12, 17):
    for col in range(1, 7):
        cell = ws_growth.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col >= 5:
            cell.number_format = '£#,##0'

# Moderate Growth
ws_growth['A19'] = "MODERATE GROWTH (25 new firms/year)"
ws_growth['A19'].font = subheader_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=21, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Moderate data
moderate_new = [25, 35, 45, 60, 75]
for i, (year, new_firms) in enumerate(zip(range(1, 6), moderate_new)):
    row = 22 + i
    ws_growth.cell(row=row, column=1, value=year)
    ws_growth.cell(row=row, column=2, value=new_firms)
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0)
        ws_growth.cell(row=row, column=4, value=f'=B{row}-C{row}')
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6/100,0)')
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}')
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5')
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12')

    for col in range(1, 7):
        cell = ws_growth.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col >= 5:
            cell.number_format = '£#,##0'

# Aggressive Growth
ws_growth['A29'] = "AGGRESSIVE GROWTH (50 new firms/year)"
ws_growth['A29'].font = subheader_font

for col, header in enumerate(growth_headers, start=1):
    cell = ws_growth.cell(row=31, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Aggressive data
aggressive_new = [50, 70, 100, 130, 150]
for i, (year, new_firms) in enumerate(zip(range(1, 6), aggressive_new)):
    row = 32 + i
    ws_growth.cell(row=row, column=1, value=year)
    ws_growth.cell(row=row, column=2, value=new_firms)
    if i == 0:
        ws_growth.cell(row=row, column=3, value=0)
        ws_growth.cell(row=row, column=4, value=f'=B{row}-C{row}')
    else:
        ws_growth.cell(row=row, column=3, value=f'=ROUND(D{row-1}*$B$6/100,0)')
        ws_growth.cell(row=row, column=4, value=f'=D{row-1}+B{row}-C{row}')
    ws_growth.cell(row=row, column=5, value=f'=D{row}*$B$5')
    ws_growth.cell(row=row, column=6, value=f'=E{row}*12')

    for col in range(1, 7):
        cell = ws_growth.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if col >= 5:
            cell.number_format = '£#,##0'

# Column widths
for col in range(1, 7):
    ws_growth.column_dimensions[get_column_letter(col)].width = 15

# ============================================
# SHEET 4: COMPETITIVE ANALYSIS
# ============================================
ws_comp = wb.create_sheet("Competitive Analysis")

ws_comp['A1'] = "COMPETITIVE ANALYSIS"
ws_comp['A1'].font = title_font
ws_comp.merge_cells('A1:E1')

ws_comp['A3'] = "UK IFA Software Market Comparison"
ws_comp['A3'].font = subheader_font

comp_headers = ['Competitor', 'Monthly Price', 'What They Offer', 'Plannetic Equivalent', 'Our Advantage']
for col, header in enumerate(comp_headers, start=1):
    cell = ws_comp.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

competitors = [
    ['Intelliflo Office', '£200-400', 'Back office, limited planning', 'Full platform', 'More features, lower price'],
    ['Voyant', '£150', 'Cash flow only', 'Cash Flow module', 'Includes CRM + compliance'],
    ['Timeline', '£100', 'Monte Carlo only', 'Monte Carlo module', 'Full assessment suite'],
    ['CashCalc', '£80', 'Cash flow only', 'Cash Flow module', 'Integrated compliance'],
    ['Salesforce', '£150+', 'Generic CRM', 'Client Hub', 'IFA-specific features'],
    ['Dynamic Planner', '£200+', 'Risk profiling + reports', 'ATR/CFL + Docs', 'Consumer Duty built-in'],
    ['FE Analytics', '£100+', 'Fund analysis', 'N/A', 'Different focus'],
    ['Defaqto Engage', '£150+', 'Research + suitability', 'Suitability module', 'Full workflow'],
]

for row_idx, row_data in enumerate(competitors, start=6):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# Tool Stack Comparison
ws_comp['A16'] = "Typical IFA Tool Stack vs Plannetic"
ws_comp['A16'].font = subheader_font

stack_headers = ['Tool Category', 'Standalone Cost', 'Plannetic', 'Savings']
for col, header in enumerate(stack_headers, start=1):
    cell = ws_comp.cell(row=18, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

stack_data = [
    ['CRM', 50, 'Included', '=B19'],
    ['Risk Profiling', 50, 'Included', '=B20'],
    ['Cash Flow', 100, 'Included', '=B21'],
    ['Monte Carlo', 100, 'Included', '=B22'],
    ['Document Generation', 50, 'Included', '=B23'],
    ['E-Signatures', 20, 'Included', '=B24'],
    ['Compliance Tracking', 50, 'Included', '=B25'],
    ['TOTAL', '=SUM(B19:B25)', '£250/mo', '=B26-250'],
]

for row_idx, row_data in enumerate(stack_data, start=19):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx == 2 or col_idx == 4:
            cell.number_format = '£#,##0'
        if row_idx == 26:  # Total row
            cell.font = Font(name='Arial', size=10, bold=True)
            cell.fill = highlight_fill

# Column widths
ws_comp.column_dimensions['A'].width = 22
ws_comp.column_dimensions['B'].width = 15
ws_comp.column_dimensions['C'].width = 25
ws_comp.column_dimensions['D'].width = 22
ws_comp.column_dimensions['E'].width = 25

# ============================================
# SHEET 5: TIER COMPARISON
# ============================================
ws_tiers = wb.create_sheet("Tier Comparison")

ws_tiers['A1'] = "PLANNETIC PRICING TIERS"
ws_tiers['A1'].font = title_font
ws_tiers.merge_cells('A1:D1')

tier_headers = ['Feature', 'Standard (£250/mo)', 'Professional (£300/mo)', 'Enterprise (Custom)']
for col, header in enumerate(tier_headers, start=1):
    cell = ws_tiers.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

features = [
    ['Client Management', '✓', '✓', '✓'],
    ['All 6 Assessment Types', '✓', '✓', '✓'],
    ['Unlimited Clients', '✓', '✓', '✓'],
    ['Document Generation', '✓', '✓', '✓'],
    ['E-Signatures', 'Fair Use', 'Unlimited', 'Unlimited'],
    ['Compliance Registers', '✓', '✓', '✓'],
    ['Consumer Duty Workflows', '✓', '✓', '✓'],
    ['AI Enhancement', '—', '✓', '✓'],
    ['Priority Support', '—', '✓', '✓'],
    ['Phone Support', '—', '✓', '✓'],
    ['White-label Portal', '—', '✓', '✓'],
    ['Advanced Analytics', '—', '✓', '✓'],
    ['API Access', '—', '—', '✓'],
    ['Custom Integrations', '—', '—', '✓'],
    ['Dedicated Account Manager', '—', '—', '✓'],
    ['Data Migration Support', 'Self-serve', 'Assisted', 'Full Service'],
    ['Onboarding', '2 calls', '4 calls', 'Unlimited'],
    ['', '', '', ''],
    ['Commitment', '2 years', '2 years', '3 years'],
    ['2-Year TCV', '=250*24', '=300*24', 'Custom'],
    ['3-Year TCV', '=250*36', '=300*36', 'Custom'],
]

for row_idx, row_data in enumerate(features, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_tiers.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')
        if row_idx % 2 == 0:
            cell.fill = alt_fill
        if value == '✓':
            cell.font = Font(name='Arial', size=10, color='059669')
        elif value == '—':
            cell.font = Font(name='Arial', size=10, color='9CA3AF')
        if row_idx in [22, 23] and col_idx in [2, 3]:
            cell.number_format = '£#,##0'

# Column widths
ws_tiers.column_dimensions['A'].width = 25
ws_tiers.column_dimensions['B'].width = 22
ws_tiers.column_dimensions['C'].width = 22
ws_tiers.column_dimensions['D'].width = 22

# ============================================
# SHEET 6: ROI CALCULATOR
# ============================================
ws_roi = wb.create_sheet("ROI Calculator")

ws_roi['A1'] = "CLIENT ROI CALCULATOR"
ws_roi['A1'].font = title_font
ws_roi.merge_cells('A1:D1')

ws_roi['A3'] = "Calculate ROI for your clients"
ws_roi['A3'].font = subheader_font

# Input section
ws_roi['A5'] = "INPUTS (edit yellow cells)"
ws_roi['A5'].font = Font(bold=True)

inputs = [
    ['Current CRM Cost (£/mo)', 50],
    ['Current Risk Tool Cost (£/mo)', 50],
    ['Current Cash Flow Tool (£/mo)', 100],
    ['Current Monte Carlo Tool (£/mo)', 100],
    ['Current Doc Gen Tool (£/mo)', 50],
    ['Current E-Sign Cost (£/mo)', 20],
    ['Current Compliance Tool (£/mo)', 50],
    ['', ''],
    ['Hours per client onboarding', 15],
    ['Hourly rate (£)', 100],
    ['New clients per month', 4],
    ['', ''],
    ['Plannetic Monthly Cost (£)', 250],
]

for row_idx, (label, value) in enumerate(inputs, start=7):
    ws_roi.cell(row=row_idx, column=1, value=label)
    cell = ws_roi.cell(row=row_idx, column=2, value=value if value != '' else None)
    if value != '' and label:
        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        cell.number_format = '£#,##0' if '£' in label or 'Cost' in label or 'rate' in label else '0'
    cell.border = thin_border

# Results section
ws_roi['A22'] = "RESULTS"
ws_roi['A22'].font = Font(bold=True)

results = [
    ['Current Total Monthly Cost', '=SUM(B7:B13)'],
    ['Monthly Software Savings', '=B24-B19'],
    ['Annual Software Savings', '=B25*12'],
    ['', ''],
    ['Time Saved per Client (hrs)', '=B15*0.6'],  # Assume 60% time savings
    ['Value of Time Saved per Client', '=B28*B16'],
    ['Monthly Time Value Saved', '=B29*B17'],
    ['Annual Time Value Saved', '=B30*12'],
    ['', ''],
    ['Total Annual Savings', '=B26+B31'],
    ['Annual Plannetic Cost', '=B19*12'],
    ['Net Annual Benefit', '=B33-B34'],
    ['ROI %', '=(B35/B34)*100'],
]

for row_idx, (label, formula) in enumerate(results, start=24):
    ws_roi.cell(row=row_idx, column=1, value=label)
    cell = ws_roi.cell(row=row_idx, column=2, value=formula if formula else None)
    cell.border = thin_border
    if formula and label:
        if 'ROI' in label:
            cell.number_format = '0.0"%"'
        else:
            cell.number_format = '£#,##0'
        if label in ['Total Annual Savings', 'Net Annual Benefit', 'ROI %']:
            cell.fill = highlight_fill
            cell.font = Font(bold=True)

# Column widths
ws_roi.column_dimensions['A'].width = 35
ws_roi.column_dimensions['B'].width = 18

# Save workbook
output_path = '/Users/adeomosanya/Documents/ifa-professional-portal/ifa-platform/Plannetic-Pricing-Analysis.xlsx'
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")
